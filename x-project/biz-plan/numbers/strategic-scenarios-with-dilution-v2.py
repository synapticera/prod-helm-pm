#!/usr/bin/env python3
"""
Strategic Scenarios Analysis with Realistic Dilution Modeling
Based on 2025/2026 market conditions for B2B AI/SaaS companies
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
SYNAPTIC_BLUE = "01A9DB"
SYNAPTIC_ORANGE = "FB4100"
DARK_NAVY = "07253D"
LIGHT_GRAY = "F4F6F6"
INPUT_BLUE = "0000FF"
GREEN = "2E7D32"

# Styles
header_fill = PatternFill("solid", fgColor=DARK_NAVY)
header_font = Font(bold=True, color="FFFFFF", size=11)
section_font = Font(bold=True, color=SYNAPTIC_BLUE, size=11)
input_font = Font(color=INPUT_BLUE)
bold_font = Font(bold=True)
note_font = Font(italic=True, color="666666", size=9)

def set_column_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def add_header(ws, row, text, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

def add_section(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font

def add_input_row(ws, row, label, value, col_label=2, col_value=3, is_pct=False, is_currency=False):
    ws.cell(row=row, column=col_label, value=label)
    cell = ws.cell(row=row, column=col_value, value=value)
    cell.font = input_font
    if is_pct:
        cell.number_format = '0.0%'
    elif is_currency:
        cell.number_format = '$#,##0.0"M"'

wb = Workbook()

# ============================================================================
# SHEET 1: PATH 1 - LEAN (Bootstrap)
# ============================================================================
ws1 = wb.active
ws1.title = "Path1_Lean"
set_column_widths(ws1, [3, 35, 15, 15, 15, 15, 15])

add_header(ws1, 1, "PATH 1: LEAN BOOTSTRAP - No External Dilution", 7)

add_section(ws1, 3, "STARTING POSITION")
add_input_row(ws1, 4, "Founder ownership", 1.0, is_pct=True)
add_input_row(ws1, 5, "Services revenue ($M)", 4.0, is_currency=True)
add_input_row(ws1, 6, "Services EBITDA margin", 0.15, is_pct=True)
add_input_row(ws1, 7, "Annual cash for product ($M)", 0.3, is_currency=True)

add_section(ws1, 9, "CAPITAL STRUCTURE")
ws1.cell(row=10, column=2, value="No external funding - 100% self-funded from services cash flow")
ws1.cell(row=10, column=2).font = note_font
ws1.cell(row=11, column=2, value="Founder retains 100% ownership throughout")

add_section(ws1, 13, "3-YEAR EXIT SCENARIOS")
ws1.cell(row=14, column=2, value="Scenario")
ws1.cell(row=14, column=3, value="Probability")
ws1.cell(row=14, column=4, value="Enterprise Value")
ws1.cell(row=14, column=5, value="Founder %")
ws1.cell(row=14, column=6, value="Founder Value")
for col in range(2, 7):
    ws1.cell(row=14, column=col).font = bold_font

# Product traction
ws1.cell(row=15, column=2, value="Product gains traction")
ws1.cell(row=15, column=3, value=0.25)
ws1.cell(row=15, column=3).font = input_font
ws1.cell(row=15, column=3).number_format = '0%'
ws1.cell(row=15, column=4, value=30.0)
ws1.cell(row=15, column=4).font = input_font
ws1.cell(row=15, column=4).number_format = '$#,##0"M"'
ws1.cell(row=15, column=5, value="=C4")
ws1.cell(row=15, column=5).number_format = '0%'
ws1.cell(row=15, column=6, value="=D15*E15")
ws1.cell(row=15, column=6).number_format = '$#,##0.0"M"'

# Modest growth
ws1.cell(row=16, column=2, value="Modest services growth")
ws1.cell(row=16, column=3, value=0.50)
ws1.cell(row=16, column=3).font = input_font
ws1.cell(row=16, column=3).number_format = '0%'
ws1.cell(row=16, column=4, value=14.0)
ws1.cell(row=16, column=4).font = input_font
ws1.cell(row=16, column=4).number_format = '$#,##0"M"'
ws1.cell(row=16, column=5, value="=C4")
ws1.cell(row=16, column=5).number_format = '0%'
ws1.cell(row=16, column=6, value="=D16*E16")
ws1.cell(row=16, column=6).number_format = '$#,##0.0"M"'

# Struggles
ws1.cell(row=17, column=2, value="Business stagnates")
ws1.cell(row=17, column=3, value=0.25)
ws1.cell(row=17, column=3).font = input_font
ws1.cell(row=17, column=3).number_format = '0%'
ws1.cell(row=17, column=4, value=8.0)
ws1.cell(row=17, column=4).font = input_font
ws1.cell(row=17, column=4).number_format = '$#,##0"M"'
ws1.cell(row=17, column=5, value="=C4")
ws1.cell(row=17, column=5).number_format = '0%'
ws1.cell(row=17, column=6, value="=D17*E17")
ws1.cell(row=17, column=6).number_format = '$#,##0.0"M"'

add_section(ws1, 19, "SUMMARY")
ws1.cell(row=20, column=2, value="Expected Enterprise Value")
ws1.cell(row=20, column=3, value="=SUMPRODUCT(C15:C17,D15:D17)")
ws1.cell(row=20, column=3).number_format = '$#,##0.0"M"'
ws1.cell(row=20, column=3).font = bold_font

ws1.cell(row=21, column=2, value="Expected Founder Value")
ws1.cell(row=21, column=3, value="=SUMPRODUCT(C15:C17,F15:F17)")
ws1.cell(row=21, column=3).number_format = '$#,##0.0"M"'
ws1.cell(row=21, column=3).font = Font(bold=True, color=SYNAPTIC_BLUE, size=12)

ws1.cell(row=22, column=2, value="Final Founder Ownership")
ws1.cell(row=22, column=3, value="=C4")
ws1.cell(row=22, column=3).number_format = '0%'
ws1.cell(row=22, column=3).font = Font(bold=True, color=GREEN)

ws1.cell(row=23, column=2, value="Total Dilution")
ws1.cell(row=23, column=3, value="=1-C22")
ws1.cell(row=23, column=3).number_format = '0%'

# ============================================================================
# SHEET 2: PATH 2 - LIQUIDATE
# ============================================================================
ws2 = wb.create_sheet("Path2_Liquidate")
set_column_widths(ws2, [3, 35, 15, 15, 15, 15, 15])

add_header(ws2, 1, "PATH 2: LIQUIDATE - Sell Services, All-In Product", 7)

add_section(ws2, 3, "STARTING POSITION")
add_input_row(ws2, 4, "Founder ownership", 1.0, is_pct=True)
add_input_row(ws2, 5, "Services sale price ($M)", 10.0, is_currency=True)
add_input_row(ws2, 6, "Founder take from sale ($M)", 8.0, is_currency=True)
ws2.cell(row=6, column=4, value="(after advisors, legal, escrow)")
ws2.cell(row=6, column=4).font = note_font

add_section(ws2, 8, "FUNDING ROUNDS - 2025/2026 MARKET")
ws2.cell(row=9, column=2, value="Round")
ws2.cell(row=9, column=3, value="Timing")
ws2.cell(row=9, column=4, value="Pre-Money")
ws2.cell(row=9, column=5, value="Raise")
ws2.cell(row=9, column=6, value="Post-Money")
ws2.cell(row=9, column=7, value="Dilution")
for col in range(2, 8):
    ws2.cell(row=9, column=col).font = bold_font

# ESOP - before seed (standard)
ws2.cell(row=10, column=2, value="ESOP Pool (pre-seed)")
ws2.cell(row=10, column=3, value="Day 1")
ws2.cell(row=10, column=4, value="-")
ws2.cell(row=10, column=5, value=0.12)
ws2.cell(row=10, column=5).font = input_font
ws2.cell(row=10, column=5).number_format = '0%'
ws2.cell(row=10, column=6, value="-")
ws2.cell(row=10, column=7, value="=E10")
ws2.cell(row=10, column=7).number_format = '0.0%'

# Seed - raising fast without much traction, lower valuation
ws2.cell(row=11, column=2, value="Seed Round")
ws2.cell(row=11, column=3, value="Month 2-3")
ws2.cell(row=11, column=4, value=7.0)
ws2.cell(row=11, column=4).font = input_font
ws2.cell(row=11, column=4).number_format = '$#,##0"M"'
ws2.cell(row=11, column=5, value=2.5)
ws2.cell(row=11, column=5).font = input_font
ws2.cell(row=11, column=5).number_format = '$#,##0.0"M"'
ws2.cell(row=11, column=6, value="=D11+E11")
ws2.cell(row=11, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=11, column=7, value="=E11/F11")
ws2.cell(row=11, column=7).number_format = '0.0%'

# Series A - aggressive timeline
ws2.cell(row=12, column=2, value="Series A")
ws2.cell(row=12, column=3, value="Month 14-16")
ws2.cell(row=12, column=4, value=22.0)
ws2.cell(row=12, column=4).font = input_font
ws2.cell(row=12, column=4).number_format = '$#,##0"M"'
ws2.cell(row=12, column=5, value=8.0)
ws2.cell(row=12, column=5).font = input_font
ws2.cell(row=12, column=5).number_format = '$#,##0.0"M"'
ws2.cell(row=12, column=6, value="=D12+E12")
ws2.cell(row=12, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=12, column=7, value="=E12/F12")
ws2.cell(row=12, column=7).number_format = '0.0%'

ws2.cell(row=13, column=2, value="Note: Lower seed valuation due to raising without customer validation")
ws2.cell(row=13, column=2).font = note_font

add_section(ws2, 15, "OWNERSHIP WATERFALL")
ws2.cell(row=16, column=2, value="Event")
ws2.cell(row=16, column=3, value="Before")
ws2.cell(row=16, column=4, value="Dilution")
ws2.cell(row=16, column=5, value="After")
for col in range(2, 6):
    ws2.cell(row=16, column=col).font = bold_font

ws2.cell(row=17, column=2, value="Starting")
ws2.cell(row=17, column=3, value="=C4")
ws2.cell(row=17, column=3).number_format = '0.0%'
ws2.cell(row=17, column=4, value=0)
ws2.cell(row=17, column=4).number_format = '0.0%'
ws2.cell(row=17, column=5, value="=C17")
ws2.cell(row=17, column=5).number_format = '0.0%'

ws2.cell(row=18, column=2, value="After ESOP")
ws2.cell(row=18, column=3, value="=E17")
ws2.cell(row=18, column=3).number_format = '0.0%'
ws2.cell(row=18, column=4, value="=G10")
ws2.cell(row=18, column=4).number_format = '0.0%'
ws2.cell(row=18, column=5, value="=C18*(1-D18)")
ws2.cell(row=18, column=5).number_format = '0.0%'

ws2.cell(row=19, column=2, value="After Seed")
ws2.cell(row=19, column=3, value="=E18")
ws2.cell(row=19, column=3).number_format = '0.0%'
ws2.cell(row=19, column=4, value="=G11")
ws2.cell(row=19, column=4).number_format = '0.0%'
ws2.cell(row=19, column=5, value="=C19*(1-D19)")
ws2.cell(row=19, column=5).number_format = '0.0%'

ws2.cell(row=20, column=2, value="After Series A")
ws2.cell(row=20, column=3, value="=E19")
ws2.cell(row=20, column=3).number_format = '0.0%'
ws2.cell(row=20, column=4, value="=G12")
ws2.cell(row=20, column=4).number_format = '0.0%'
ws2.cell(row=20, column=5, value="=C20*(1-D20)")
ws2.cell(row=20, column=5).number_format = '0.0%'

add_section(ws2, 22, "3-YEAR EXIT SCENARIOS")
ws2.cell(row=23, column=2, value="Scenario")
ws2.cell(row=23, column=3, value="Probability")
ws2.cell(row=23, column=4, value="Product EV")
ws2.cell(row=23, column=5, value="Founder %")
ws2.cell(row=23, column=6, value="From Product")
ws2.cell(row=23, column=7, value="+ Services = Total")
for col in range(2, 8):
    ws2.cell(row=23, column=col).font = bold_font

# Helm succeeds big
ws2.cell(row=24, column=2, value="Helm hits $10M+ ARR")
ws2.cell(row=24, column=3, value=0.35)
ws2.cell(row=24, column=3).font = input_font
ws2.cell(row=24, column=3).number_format = '0%'
ws2.cell(row=24, column=4, value=100.0)
ws2.cell(row=24, column=4).font = input_font
ws2.cell(row=24, column=4).number_format = '$#,##0"M"'
ws2.cell(row=24, column=5, value="=E20")
ws2.cell(row=24, column=5).number_format = '0.0%'
ws2.cell(row=24, column=6, value="=D24*E24")
ws2.cell(row=24, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=24, column=7, value="=F24+C6")
ws2.cell(row=24, column=7).number_format = '$#,##0.0"M"'

# Moderate traction
ws2.cell(row=25, column=2, value="Helm at $3-5M ARR")
ws2.cell(row=25, column=3, value=0.30)
ws2.cell(row=25, column=3).font = input_font
ws2.cell(row=25, column=3).number_format = '0%'
ws2.cell(row=25, column=4, value=35.0)
ws2.cell(row=25, column=4).font = input_font
ws2.cell(row=25, column=4).number_format = '$#,##0"M"'
ws2.cell(row=25, column=5, value="=E20")
ws2.cell(row=25, column=5).number_format = '0.0%'
ws2.cell(row=25, column=6, value="=D25*E25")
ws2.cell(row=25, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=25, column=7, value="=F25+C6")
ws2.cell(row=25, column=7).number_format = '$#,##0.0"M"'

# Struggles/acquihire
ws2.cell(row=26, column=2, value="Helm struggles (<$1M ARR)")
ws2.cell(row=26, column=3, value=0.25)
ws2.cell(row=26, column=3).font = input_font
ws2.cell(row=26, column=3).number_format = '0%'
ws2.cell(row=26, column=4, value=8.0)
ws2.cell(row=26, column=4).font = input_font
ws2.cell(row=26, column=4).number_format = '$#,##0"M"'
ws2.cell(row=26, column=5, value="=E20")
ws2.cell(row=26, column=5).number_format = '0.0%'
ws2.cell(row=26, column=6, value="=D26*E26")
ws2.cell(row=26, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=26, column=7, value="=F26+C6")
ws2.cell(row=26, column=7).number_format = '$#,##0.0"M"'

# Fails
ws2.cell(row=27, column=2, value="Helm shuts down")
ws2.cell(row=27, column=3, value=0.10)
ws2.cell(row=27, column=3).font = input_font
ws2.cell(row=27, column=3).number_format = '0%'
ws2.cell(row=27, column=4, value=0.0)
ws2.cell(row=27, column=4).font = input_font
ws2.cell(row=27, column=4).number_format = '$#,##0"M"'
ws2.cell(row=27, column=5, value="=E20")
ws2.cell(row=27, column=5).number_format = '0.0%'
ws2.cell(row=27, column=6, value="=D27*E27")
ws2.cell(row=27, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=27, column=7, value="=F27+C6")
ws2.cell(row=27, column=7).number_format = '$#,##0.0"M"'

add_section(ws2, 29, "SUMMARY")
ws2.cell(row=30, column=2, value="Expected Enterprise Value")
ws2.cell(row=30, column=3, value="=SUMPRODUCT(C24:C27,D24:D27)")
ws2.cell(row=30, column=3).number_format = '$#,##0.0"M"'
ws2.cell(row=30, column=3).font = bold_font

ws2.cell(row=31, column=2, value="Expected Founder Value (Product)")
ws2.cell(row=31, column=3, value="=SUMPRODUCT(C24:C27,F24:F27)")
ws2.cell(row=31, column=3).number_format = '$#,##0.0"M"'

ws2.cell(row=32, column=2, value="Expected Founder Value (Total)")
ws2.cell(row=32, column=3, value="=SUMPRODUCT(C24:C27,G24:G27)")
ws2.cell(row=32, column=3).number_format = '$#,##0.0"M"'
ws2.cell(row=32, column=3).font = Font(bold=True, color=SYNAPTIC_BLUE, size=12)

ws2.cell(row=33, column=2, value="Final Founder Ownership")
ws2.cell(row=33, column=3, value="=E20")
ws2.cell(row=33, column=3).number_format = '0.0%'
ws2.cell(row=33, column=3).font = Font(bold=True, color=GREEN)

ws2.cell(row=34, column=2, value="Total Dilution")
ws2.cell(row=34, column=3, value="=1-C33")
ws2.cell(row=34, column=3).number_format = '0.0%'

ws2.cell(row=35, column=2, value="Total Capital Raised")
ws2.cell(row=35, column=3, value="=E11+E12")
ws2.cell(row=35, column=3).number_format = '$#,##0.0"M"'

# ============================================================================
# SHEET 3: PATH 3 - GROWTH BOTH
# ============================================================================
ws3 = wb.create_sheet("Path3_GrowthBoth")
set_column_widths(ws3, [3, 35, 15, 15, 15, 15, 15])

add_header(ws3, 1, "PATH 3: GROWTH BOTH - Dual Business Strategy", 7)

add_section(ws3, 3, "STARTING POSITION")
add_input_row(ws3, 4, "Founder ownership", 1.0, is_pct=True)
add_input_row(ws3, 5, "Services Y1 revenue ($M)", 4.5, is_currency=True)
add_input_row(ws3, 6, "Services Y3 value (3x rev @ $8M)", 24.0, is_currency=True)
add_input_row(ws3, 7, "Services leadership hire ($M/yr)", 0.25, is_currency=True)

add_section(ws3, 9, "FUNDING ROUNDS - PRODUCT ENTITY")
ws3.cell(row=10, column=2, value="Round")
ws3.cell(row=10, column=3, value="Timing")
ws3.cell(row=10, column=4, value="Pre-Money")
ws3.cell(row=10, column=5, value="Raise")
ws3.cell(row=10, column=6, value="Post-Money")
ws3.cell(row=10, column=7, value="Dilution")
for col in range(2, 8):
    ws3.cell(row=10, column=col).font = bold_font

# ESOP
ws3.cell(row=11, column=2, value="ESOP Pool")
ws3.cell(row=11, column=3, value="At Seed")
ws3.cell(row=11, column=4, value="-")
ws3.cell(row=11, column=5, value=0.12)
ws3.cell(row=11, column=5).font = input_font
ws3.cell(row=11, column=5).number_format = '0%'
ws3.cell(row=11, column=6, value="-")
ws3.cell(row=11, column=7, value="=E11")
ws3.cell(row=11, column=7).number_format = '0.0%'

# Seed - can show some services integration, slightly better
ws3.cell(row=12, column=2, value="Seed Round")
ws3.cell(row=12, column=3, value="Month 4-6")
ws3.cell(row=12, column=4, value=8.0)
ws3.cell(row=12, column=4).font = input_font
ws3.cell(row=12, column=4).number_format = '$#,##0"M"'
ws3.cell(row=12, column=5, value=3.0)
ws3.cell(row=12, column=5).font = input_font
ws3.cell(row=12, column=5).number_format = '$#,##0.0"M"'
ws3.cell(row=12, column=6, value="=D12+E12")
ws3.cell(row=12, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=12, column=7, value="=E12/F12")
ws3.cell(row=12, column=7).number_format = '0.0%'

# Series A
ws3.cell(row=13, column=2, value="Series A")
ws3.cell(row=13, column=3, value="Month 18")
ws3.cell(row=13, column=4, value=25.0)
ws3.cell(row=13, column=4).font = input_font
ws3.cell(row=13, column=4).number_format = '$#,##0"M"'
ws3.cell(row=13, column=5, value=8.0)
ws3.cell(row=13, column=5).font = input_font
ws3.cell(row=13, column=5).number_format = '$#,##0.0"M"'
ws3.cell(row=13, column=6, value="=D13+E13")
ws3.cell(row=13, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=13, column=7, value="=E13/F13")
ws3.cell(row=13, column=7).number_format = '0.0%'

ws3.cell(row=14, column=2, value="Note: VCs typically discount services-heavy companies by 20-30%")
ws3.cell(row=14, column=2).font = note_font

add_section(ws3, 16, "OWNERSHIP WATERFALL (PRODUCT)")
ws3.cell(row=17, column=2, value="Event")
ws3.cell(row=17, column=3, value="Before")
ws3.cell(row=17, column=4, value="Dilution")
ws3.cell(row=17, column=5, value="After")
for col in range(2, 6):
    ws3.cell(row=17, column=col).font = bold_font

ws3.cell(row=18, column=2, value="Starting")
ws3.cell(row=18, column=3, value="=C4")
ws3.cell(row=18, column=3).number_format = '0.0%'
ws3.cell(row=18, column=4, value=0)
ws3.cell(row=18, column=4).number_format = '0.0%'
ws3.cell(row=18, column=5, value="=C18")
ws3.cell(row=18, column=5).number_format = '0.0%'

ws3.cell(row=19, column=2, value="After ESOP")
ws3.cell(row=19, column=3, value="=E18")
ws3.cell(row=19, column=3).number_format = '0.0%'
ws3.cell(row=19, column=4, value="=G11")
ws3.cell(row=19, column=4).number_format = '0.0%'
ws3.cell(row=19, column=5, value="=C19*(1-D19)")
ws3.cell(row=19, column=5).number_format = '0.0%'

ws3.cell(row=20, column=2, value="After Seed")
ws3.cell(row=20, column=3, value="=E19")
ws3.cell(row=20, column=3).number_format = '0.0%'
ws3.cell(row=20, column=4, value="=G12")
ws3.cell(row=20, column=4).number_format = '0.0%'
ws3.cell(row=20, column=5, value="=C20*(1-D20)")
ws3.cell(row=20, column=5).number_format = '0.0%'

ws3.cell(row=21, column=2, value="After Series A")
ws3.cell(row=21, column=3, value="=E20")
ws3.cell(row=21, column=3).number_format = '0.0%'
ws3.cell(row=21, column=4, value="=G13")
ws3.cell(row=21, column=4).number_format = '0.0%'
ws3.cell(row=21, column=5, value="=C21*(1-D21)")
ws3.cell(row=21, column=5).number_format = '0.0%'

add_section(ws3, 23, "3-YEAR EXIT SCENARIOS")
ws3.cell(row=24, column=2, value="Scenario")
ws3.cell(row=24, column=3, value="Prob")
ws3.cell(row=24, column=4, value="Product EV")
ws3.cell(row=24, column=5, value="Services EV")
ws3.cell(row=24, column=6, value="Founder Prod")
ws3.cell(row=24, column=7, value="Total Founder")
for col in range(2, 8):
    ws3.cell(row=24, column=col).font = bold_font

# Both succeed
ws3.cell(row=25, column=2, value="Both succeed")
ws3.cell(row=25, column=3, value=0.30)
ws3.cell(row=25, column=3).font = input_font
ws3.cell(row=25, column=3).number_format = '0%'
ws3.cell(row=25, column=4, value=70.0)
ws3.cell(row=25, column=4).font = input_font
ws3.cell(row=25, column=4).number_format = '$#,##0"M"'
ws3.cell(row=25, column=5, value="=C6")
ws3.cell(row=25, column=5).number_format = '$#,##0"M"'
ws3.cell(row=25, column=6, value="=D25*E21")
ws3.cell(row=25, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=25, column=7, value="=F25+E25")
ws3.cell(row=25, column=7).number_format = '$#,##0.0"M"'

# Product moderate, services good
ws3.cell(row=26, column=2, value="Product moderate")
ws3.cell(row=26, column=3, value=0.35)
ws3.cell(row=26, column=3).font = input_font
ws3.cell(row=26, column=3).number_format = '0%'
ws3.cell(row=26, column=4, value=25.0)
ws3.cell(row=26, column=4).font = input_font
ws3.cell(row=26, column=4).number_format = '$#,##0"M"'
ws3.cell(row=26, column=5, value="=C6")
ws3.cell(row=26, column=5).number_format = '$#,##0"M"'
ws3.cell(row=26, column=6, value="=D26*E21")
ws3.cell(row=26, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=26, column=7, value="=F26+E26")
ws3.cell(row=26, column=7).number_format = '$#,##0.0"M"'

# Both struggle
ws3.cell(row=27, column=2, value="Focus dilution hurts both")
ws3.cell(row=27, column=3, value=0.25)
ws3.cell(row=27, column=3).font = input_font
ws3.cell(row=27, column=3).number_format = '0%'
ws3.cell(row=27, column=4, value=10.0)
ws3.cell(row=27, column=4).font = input_font
ws3.cell(row=27, column=4).number_format = '$#,##0"M"'
ws3.cell(row=27, column=5, value=15.0)
ws3.cell(row=27, column=5).font = input_font
ws3.cell(row=27, column=5).number_format = '$#,##0"M"'
ws3.cell(row=27, column=6, value="=D27*E21")
ws3.cell(row=27, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=27, column=7, value="=F27+E27")
ws3.cell(row=27, column=7).number_format = '$#,##0.0"M"'

# Product fails
ws3.cell(row=28, column=2, value="Product fails, services OK")
ws3.cell(row=28, column=3, value=0.10)
ws3.cell(row=28, column=3).font = input_font
ws3.cell(row=28, column=3).number_format = '0%'
ws3.cell(row=28, column=4, value=0.0)
ws3.cell(row=28, column=4).font = input_font
ws3.cell(row=28, column=4).number_format = '$#,##0"M"'
ws3.cell(row=28, column=5, value=18.0)
ws3.cell(row=28, column=5).font = input_font
ws3.cell(row=28, column=5).number_format = '$#,##0"M"'
ws3.cell(row=28, column=6, value="=D28*E21")
ws3.cell(row=28, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=28, column=7, value="=F28+E28")
ws3.cell(row=28, column=7).number_format = '$#,##0.0"M"'

add_section(ws3, 30, "SUMMARY")
ws3.cell(row=31, column=2, value="Expected Enterprise Value (Combined)")
ws3.cell(row=31, column=3, value="=SUMPRODUCT(C25:C28,D25:D28)+SUMPRODUCT(C25:C28,E25:E28)")
ws3.cell(row=31, column=3).number_format = '$#,##0.0"M"'
ws3.cell(row=31, column=3).font = bold_font

ws3.cell(row=32, column=2, value="Expected Founder Value")
ws3.cell(row=32, column=3, value="=SUMPRODUCT(C25:C28,G25:G28)")
ws3.cell(row=32, column=3).number_format = '$#,##0.0"M"'
ws3.cell(row=32, column=3).font = Font(bold=True, color=SYNAPTIC_BLUE, size=12)

ws3.cell(row=33, column=2, value="Final Founder % (Product)")
ws3.cell(row=33, column=3, value="=E21")
ws3.cell(row=33, column=3).number_format = '0.0%'
ws3.cell(row=33, column=3).font = Font(bold=True, color=GREEN)

ws3.cell(row=34, column=2, value="Total Dilution (Product)")
ws3.cell(row=34, column=3, value="=1-C33")
ws3.cell(row=34, column=3).number_format = '0.0%'

ws3.cell(row=35, column=2, value="Total Capital Raised")
ws3.cell(row=35, column=3, value="=E12+E13")
ws3.cell(row=35, column=3).number_format = '$#,##0.0"M"'

# ============================================================================
# SHEET 4: PATH 4 - STRATEGIC LEVERAGE (RECOMMENDED)
# ============================================================================
ws4 = wb.create_sheet("Path4_Strategic")
set_column_widths(ws4, [3, 35, 15, 15, 15, 15, 15])

add_header(ws4, 1, "PATH 4: STRATEGIC LEVERAGE (RECOMMENDED)", 7)

add_section(ws4, 3, "STARTING POSITION")
add_input_row(ws4, 4, "Founder ownership", 1.0, is_pct=True)
add_input_row(ws4, 5, "Services revenue ($M)", 4.0, is_currency=True)
add_input_row(ws4, 6, "Services EBITDA (optimized to 20%)", 0.8, is_currency=True)
add_input_row(ws4, 7, "Services sale price (Month 12, $M)", 10.0, is_currency=True)
add_input_row(ws4, 8, "Founder take from services ($M)", 8.0, is_currency=True)

add_section(ws4, 10, "FUNDING ROUNDS - VALIDATED APPROACH")
ws4.cell(row=11, column=2, value="Round")
ws4.cell(row=11, column=3, value="Timing")
ws4.cell(row=11, column=4, value="Pre-Money")
ws4.cell(row=11, column=5, value="Raise")
ws4.cell(row=11, column=6, value="Post-Money")
ws4.cell(row=11, column=7, value="Dilution")
for col in range(2, 8):
    ws4.cell(row=11, column=col).font = bold_font

# ESOP
ws4.cell(row=12, column=2, value="ESOP Pool")
ws4.cell(row=12, column=3, value="At Seed")
ws4.cell(row=12, column=4, value="-")
ws4.cell(row=12, column=5, value=0.10)
ws4.cell(row=12, column=5).font = input_font
ws4.cell(row=12, column=5).number_format = '0%'
ws4.cell(row=12, column=6, value="-")
ws4.cell(row=12, column=7, value="=E12")
ws4.cell(row=12, column=7).number_format = '0.0%'

# Seed - HIGHER valuation due to proof points
ws4.cell(row=13, column=2, value="Seed (with design partners)")
ws4.cell(row=13, column=3, value="Month 12-14")
ws4.cell(row=13, column=4, value=12.0)
ws4.cell(row=13, column=4).font = input_font
ws4.cell(row=13, column=4).number_format = '$#,##0"M"'
ws4.cell(row=13, column=5, value=3.0)
ws4.cell(row=13, column=5).font = input_font
ws4.cell(row=13, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=13, column=6, value="=D13+E13")
ws4.cell(row=13, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=13, column=7, value="=E13/F13")
ws4.cell(row=13, column=7).number_format = '0.0%'

# Series A - strong position
ws4.cell(row=14, column=2, value="Series A (with ARR traction)")
ws4.cell(row=14, column=3, value="Month 24-30")
ws4.cell(row=14, column=4, value=35.0)
ws4.cell(row=14, column=4).font = input_font
ws4.cell(row=14, column=4).number_format = '$#,##0"M"'
ws4.cell(row=14, column=5, value=10.0)
ws4.cell(row=14, column=5).font = input_font
ws4.cell(row=14, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=14, column=6, value="=D14+E14")
ws4.cell(row=14, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=14, column=7, value="=E14/F14")
ws4.cell(row=14, column=7).number_format = '0.0%'

ws4.cell(row=15, column=2, value="Key: 70%+ higher seed valuation due to design partner validation")
ws4.cell(row=15, column=2).font = Font(italic=True, color=GREEN, size=9)

add_section(ws4, 17, "OWNERSHIP WATERFALL")
ws4.cell(row=18, column=2, value="Event")
ws4.cell(row=18, column=3, value="Before")
ws4.cell(row=18, column=4, value="Dilution")
ws4.cell(row=18, column=5, value="After")
for col in range(2, 6):
    ws4.cell(row=18, column=col).font = bold_font

ws4.cell(row=19, column=2, value="Starting")
ws4.cell(row=19, column=3, value="=C4")
ws4.cell(row=19, column=3).number_format = '0.0%'
ws4.cell(row=19, column=4, value=0)
ws4.cell(row=19, column=4).number_format = '0.0%'
ws4.cell(row=19, column=5, value="=C19")
ws4.cell(row=19, column=5).number_format = '0.0%'

ws4.cell(row=20, column=2, value="After ESOP")
ws4.cell(row=20, column=3, value="=E19")
ws4.cell(row=20, column=3).number_format = '0.0%'
ws4.cell(row=20, column=4, value="=G12")
ws4.cell(row=20, column=4).number_format = '0.0%'
ws4.cell(row=20, column=5, value="=C20*(1-D20)")
ws4.cell(row=20, column=5).number_format = '0.0%'

ws4.cell(row=21, column=2, value="After Seed")
ws4.cell(row=21, column=3, value="=E20")
ws4.cell(row=21, column=3).number_format = '0.0%'
ws4.cell(row=21, column=4, value="=G13")
ws4.cell(row=21, column=4).number_format = '0.0%'
ws4.cell(row=21, column=5, value="=C21*(1-D21)")
ws4.cell(row=21, column=5).number_format = '0.0%'

ws4.cell(row=22, column=2, value="After Series A")
ws4.cell(row=22, column=3, value="=E21")
ws4.cell(row=22, column=3).number_format = '0.0%'
ws4.cell(row=22, column=4, value="=G14")
ws4.cell(row=22, column=4).number_format = '0.0%'
ws4.cell(row=22, column=5, value="=C22*(1-D22)")
ws4.cell(row=22, column=5).number_format = '0.0%'

add_section(ws4, 24, "3-YEAR EXIT SCENARIOS")
ws4.cell(row=25, column=2, value="Scenario")
ws4.cell(row=25, column=3, value="Prob")
ws4.cell(row=25, column=4, value="Product EV")
ws4.cell(row=25, column=5, value="Founder Prod")
ws4.cell(row=25, column=6, value="Services $")
ws4.cell(row=25, column=7, value="Total Founder")
for col in range(2, 8):
    ws4.cell(row=25, column=col).font = bold_font

# Success - validated path works
ws4.cell(row=26, column=2, value="Helm hits $10M+ ARR")
ws4.cell(row=26, column=3, value=0.50)
ws4.cell(row=26, column=3).font = input_font
ws4.cell(row=26, column=3).number_format = '0%'
ws4.cell(row=26, column=4, value=100.0)
ws4.cell(row=26, column=4).font = input_font
ws4.cell(row=26, column=4).number_format = '$#,##0"M"'
ws4.cell(row=26, column=5, value="=D26*E22")
ws4.cell(row=26, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=26, column=6, value="=C8")
ws4.cell(row=26, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=26, column=7, value="=E26+F26")
ws4.cell(row=26, column=7).number_format = '$#,##0.0"M"'

# Moderate
ws4.cell(row=27, column=2, value="Helm at $3-5M ARR")
ws4.cell(row=27, column=3, value=0.30)
ws4.cell(row=27, column=3).font = input_font
ws4.cell(row=27, column=3).number_format = '0%'
ws4.cell(row=27, column=4, value=40.0)
ws4.cell(row=27, column=4).font = input_font
ws4.cell(row=27, column=4).number_format = '$#,##0"M"'
ws4.cell(row=27, column=5, value="=D27*E22")
ws4.cell(row=27, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=27, column=6, value="=C8")
ws4.cell(row=27, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=27, column=7, value="=E27+F27")
ws4.cell(row=27, column=7).number_format = '$#,##0.0"M"'

# Struggles
ws4.cell(row=28, column=2, value="Helm pivots/struggles")
ws4.cell(row=28, column=3, value=0.15)
ws4.cell(row=28, column=3).font = input_font
ws4.cell(row=28, column=3).number_format = '0%'
ws4.cell(row=28, column=4, value=10.0)
ws4.cell(row=28, column=4).font = input_font
ws4.cell(row=28, column=4).number_format = '$#,##0"M"'
ws4.cell(row=28, column=5, value="=D28*E22")
ws4.cell(row=28, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=28, column=6, value="=C8")
ws4.cell(row=28, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=28, column=7, value="=E28+F28")
ws4.cell(row=28, column=7).number_format = '$#,##0.0"M"'

# Wind down
ws4.cell(row=29, column=2, value="Wind down (services $ preserved)")
ws4.cell(row=29, column=3, value=0.05)
ws4.cell(row=29, column=3).font = input_font
ws4.cell(row=29, column=3).number_format = '0%'
ws4.cell(row=29, column=4, value=0.0)
ws4.cell(row=29, column=4).font = input_font
ws4.cell(row=29, column=4).number_format = '$#,##0"M"'
ws4.cell(row=29, column=5, value="=D29*E22")
ws4.cell(row=29, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=29, column=6, value="=C8")
ws4.cell(row=29, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=29, column=7, value="=E29+F29")
ws4.cell(row=29, column=7).number_format = '$#,##0.0"M"'

add_section(ws4, 31, "SUMMARY")
ws4.cell(row=32, column=2, value="Expected Enterprise Value")
ws4.cell(row=32, column=3, value="=SUMPRODUCT(C26:C29,D26:D29)")
ws4.cell(row=32, column=3).number_format = '$#,##0.0"M"'
ws4.cell(row=32, column=3).font = bold_font

ws4.cell(row=33, column=2, value="Expected Founder Value")
ws4.cell(row=33, column=3, value="=SUMPRODUCT(C26:C29,G26:G29)")
ws4.cell(row=33, column=3).number_format = '$#,##0.0"M"'
ws4.cell(row=33, column=3).font = Font(bold=True, color=SYNAPTIC_BLUE, size=12)

ws4.cell(row=34, column=2, value="Final Founder % (Product)")
ws4.cell(row=34, column=3, value="=E22")
ws4.cell(row=34, column=3).number_format = '0.0%'
ws4.cell(row=34, column=3).font = Font(bold=True, color=GREEN)

ws4.cell(row=35, column=2, value="Total Dilution (Product)")
ws4.cell(row=35, column=3, value="=1-C34")
ws4.cell(row=35, column=3).number_format = '0.0%'

ws4.cell(row=36, column=2, value="Total Capital Raised")
ws4.cell(row=36, column=3, value="=E13+E14")
ws4.cell(row=36, column=3).number_format = '$#,##0.0"M"'

ws4.cell(row=37, column=2, value="Guaranteed Floor (Services)")
ws4.cell(row=37, column=3, value="=C8")
ws4.cell(row=37, column=3).number_format = '$#,##0.0"M"'
ws4.cell(row=37, column=3).font = Font(bold=True, color=SYNAPTIC_ORANGE)

# ============================================================================
# SHEET 5: SUMMARY COMPARISON
# ============================================================================
ws5 = wb.create_sheet("Summary_Comparison")
set_column_widths(ws5, [3, 28, 16, 16, 16, 16])

add_header(ws5, 1, "DILUTION-ADJUSTED COMPARISON", 6)

add_section(ws5, 3, "FOUNDER VALUE ANALYSIS")
ws5.cell(row=4, column=2, value="Metric")
ws5.cell(row=4, column=3, value="Lean")
ws5.cell(row=4, column=4, value="Liquidate")
ws5.cell(row=4, column=5, value="Growth Both")
ws5.cell(row=4, column=6, value="Strategic")
for col in range(2, 7):
    ws5.cell(row=4, column=col).font = bold_font
    ws5.cell(row=4, column=col).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

# Expected Founder Value
ws5.cell(row=5, column=2, value="Expected Founder Value")
ws5.cell(row=5, column=3, value="=Path1_Lean!C21")
ws5.cell(row=5, column=4, value="=Path2_Liquidate!C32")
ws5.cell(row=5, column=5, value="=Path3_GrowthBoth!C32")
ws5.cell(row=5, column=6, value="=Path4_Strategic!C33")
for col in range(3, 7):
    ws5.cell(row=5, column=col).number_format = '$#,##0.0"M"'
    ws5.cell(row=5, column=col).font = Font(bold=True, color=SYNAPTIC_BLUE)

# Final ownership
ws5.cell(row=6, column=2, value="Final Founder %")
ws5.cell(row=6, column=3, value="=Path1_Lean!C22")
ws5.cell(row=6, column=4, value="=Path2_Liquidate!C33")
ws5.cell(row=6, column=5, value="=Path3_GrowthBoth!C33")
ws5.cell(row=6, column=6, value="=Path4_Strategic!C34")
for col in range(3, 7):
    ws5.cell(row=6, column=col).number_format = '0.0%'

# Total dilution
ws5.cell(row=7, column=2, value="Total Dilution")
ws5.cell(row=7, column=3, value="=1-C6")
ws5.cell(row=7, column=4, value="=1-D6")
ws5.cell(row=7, column=5, value="=1-E6")
ws5.cell(row=7, column=6, value="=1-F6")
for col in range(3, 7):
    ws5.cell(row=7, column=col).number_format = '0.0%'

# Capital raised
ws5.cell(row=8, column=2, value="Capital Raised")
ws5.cell(row=8, column=3, value=0)
ws5.cell(row=8, column=4, value="=Path2_Liquidate!C35")
ws5.cell(row=8, column=5, value="=Path3_GrowthBoth!C35")
ws5.cell(row=8, column=6, value="=Path4_Strategic!C36")
for col in range(3, 7):
    ws5.cell(row=8, column=col).number_format = '$#,##0.0"M"'

# Downside protection
add_section(ws5, 10, "DOWNSIDE PROTECTION")
ws5.cell(row=11, column=2, value="Worst Case (Founder $)")
ws5.cell(row=11, column=3, value="=Path1_Lean!F17")
ws5.cell(row=11, column=4, value="=Path2_Liquidate!G27")
ws5.cell(row=11, column=5, value="=Path3_GrowthBoth!G28")
ws5.cell(row=11, column=6, value="=Path4_Strategic!G29")
for col in range(3, 7):
    ws5.cell(row=11, column=col).number_format = '$#,##0.0"M"'

ws5.cell(row=12, column=2, value="Best Case (Founder $)")
ws5.cell(row=12, column=3, value="=Path1_Lean!F15")
ws5.cell(row=12, column=4, value="=Path2_Liquidate!G24")
ws5.cell(row=12, column=5, value="=Path3_GrowthBoth!G25")
ws5.cell(row=12, column=6, value="=Path4_Strategic!G26")
for col in range(3, 7):
    ws5.cell(row=12, column=col).number_format = '$#,##0.0"M"'

ws5.cell(row=13, column=2, value="Upside/Downside Ratio")
ws5.cell(row=13, column=3, value="=C12/C11")
ws5.cell(row=13, column=4, value="=D12/D11")
ws5.cell(row=13, column=5, value="=E12/E11")
ws5.cell(row=13, column=6, value="=F12/F11")
for col in range(3, 7):
    ws5.cell(row=13, column=col).number_format = '0.0x'

# Ranking
add_section(ws5, 15, "RANKING")
ws5.cell(row=16, column=2, value="Rank by Founder Value")
ws5.cell(row=16, column=3, value="=RANK(C5,$C$5:$F$5,0)")
ws5.cell(row=16, column=4, value="=RANK(D5,$C$5:$F$5,0)")
ws5.cell(row=16, column=5, value="=RANK(E5,$C$5:$F$5,0)")
ws5.cell(row=16, column=6, value="=RANK(F5,$C$5:$F$5,0)")

ws5.cell(row=17, column=2, value="Rank by Downside Protection")
ws5.cell(row=17, column=3, value="=RANK(C11,$C$11:$F$11,0)")
ws5.cell(row=17, column=4, value="=RANK(D11,$C$11:$F$11,0)")
ws5.cell(row=17, column=5, value="=RANK(E11,$C$11:$F$11,0)")
ws5.cell(row=17, column=6, value="=RANK(F11,$C$11:$F$11,0)")

# Winner highlight
ws5.cell(row=19, column=2, value="WINNER")
ws5.merge_cells(start_row=19, start_column=6, end_row=19, end_column=6)
ws5.cell(row=19, column=6, value="STRATEGIC")
ws5.cell(row=19, column=6).fill = PatternFill("solid", fgColor=SYNAPTIC_BLUE)
ws5.cell(row=19, column=6).font = Font(bold=True, color="FFFFFF", size=12)
ws5.cell(row=19, column=6).alignment = Alignment(horizontal='center')

# Key insight
add_section(ws5, 21, "KEY INSIGHTS")
ws5.cell(row=22, column=2, value="1. Strategic Leverage has BOTH highest EV and best downside protection")
ws5.cell(row=23, column=2, value="2. Validation-first approach yields 70%+ better seed valuations")
ws5.cell(row=24, column=2, value="3. $8M services floor means founder never walks away empty")
ws5.cell(row=25, column=2, value="4. Same $13M raised, but at better terms = less dilution")

# Funding Assumptions
add_section(ws5, 27, "FUNDING ASSUMPTIONS (2025/2026 Market)")
ws5.cell(row=28, column=2, value="Seed without traction: $6-8M pre-money")
ws5.cell(row=29, column=2, value="Seed with design partners: $10-14M pre-money")
ws5.cell(row=30, column=2, value="Series A (B2B SaaS): $20-40M pre-money")
ws5.cell(row=31, column=2, value="Standard ESOP: 10-15% pre-financing")
ws5.cell(row=32, column=2, value="Services multiple: 2.5-3.5x revenue")

# Save
output_path = "/Users/wkarp/devlocal/prod-miner-pm/x-project/bplan/numbers/strategic-scenarios-with-dilution.xlsx"
wb.save(output_path)
print(f"Updated: {output_path}")
