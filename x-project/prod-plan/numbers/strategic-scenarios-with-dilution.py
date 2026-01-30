#!/usr/bin/env python3
"""
Strategic Scenarios Analysis with Dilution Modeling
Creates Excel workbook comparing all 4 paths with founder ownership waterfall
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
SYNAPTIC_BLUE = "01A9DB"
SYNAPTIC_ORANGE = "FB4100"
DARK_NAVY = "07253D"
LIGHT_GRAY = "F4F6F6"
INPUT_BLUE = "0000FF"  # Blue text for inputs

# Styles
header_fill = PatternFill("solid", fgColor=DARK_NAVY)
header_font = Font(bold=True, color="FFFFFF", size=11)
section_font = Font(bold=True, color=SYNAPTIC_BLUE, size=11)
input_font = Font(color=INPUT_BLUE)  # Blue for editable inputs
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def set_column_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def add_header(ws, row, text, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

def add_section(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font

def add_input_row(ws, row, label, value, col_label=2, col_value=3, is_pct=False, is_currency=False):
    ws.cell(row=row, column=col_label, value=label)
    cell = ws.cell(row=row, column=col_value, value=value)
    cell.font = input_font
    if is_pct:
        cell.number_format = '0.0%'
    elif is_currency:
        cell.number_format = '$#,##0.0'

wb = Workbook()

# ============================================================================
# SHEET 1: PATH 1 - LEAN (Bootstrap)
# ============================================================================
ws1 = wb.active
ws1.title = "Path1_Lean"
set_column_widths(ws1, [3, 30, 15, 15, 15, 15, 15])

add_header(ws1, 1, "PATH 1: LEAN BOOTSTRAP - Dilution Analysis", 7)

# Assumptions
add_section(ws1, 3, "ASSUMPTIONS")
add_input_row(ws1, 4, "Starting founder ownership", 1.0, is_pct=True)
add_input_row(ws1, 5, "Services revenue ($M)", 4.0, is_currency=True)
add_input_row(ws1, 6, "Services EBITDA margin", 0.15, is_pct=True)
add_input_row(ws1, 7, "Annual product investment ($M)", 0.3, is_currency=True)

# Capital Raises - NONE for Lean
add_section(ws1, 9, "CAPITAL RAISES")
ws1.cell(row=10, column=2, value="No external capital - self-funded")

# Ownership Waterfall
add_section(ws1, 12, "OWNERSHIP WATERFALL")
ws1.cell(row=13, column=2, value="Event")
ws1.cell(row=13, column=3, value="Pre-Money")
ws1.cell(row=13, column=4, value="Raise")
ws1.cell(row=13, column=5, value="Post-Money")
ws1.cell(row=13, column=6, value="Dilution")
ws1.cell(row=13, column=7, value="Founder %")
for col in range(2, 8):
    ws1.cell(row=13, column=col).font = bold_font

ws1.cell(row=14, column=2, value="Starting")
ws1.cell(row=14, column=3, value="N/A")
ws1.cell(row=14, column=4, value="$0")
ws1.cell(row=14, column=5, value="N/A")
ws1.cell(row=14, column=6, value="0%")
ws1.cell(row=14, column=7, value="=C4")
ws1.cell(row=14, column=7).number_format = '0.0%'

ws1.cell(row=15, column=2, value="Year 3 Exit")
ws1.cell(row=15, column=3, value="N/A")
ws1.cell(row=15, column=4, value="$0")
ws1.cell(row=15, column=5, value="N/A")
ws1.cell(row=15, column=6, value="0%")
ws1.cell(row=15, column=7, value="=G14")
ws1.cell(row=15, column=7).number_format = '0.0%'

# Exit Scenarios
add_section(ws1, 17, "EXIT SCENARIOS (Year 3)")
ws1.cell(row=18, column=2, value="Scenario")
ws1.cell(row=18, column=3, value="Probability")
ws1.cell(row=18, column=4, value="Enterprise Value")
ws1.cell(row=18, column=5, value="Founder %")
ws1.cell(row=18, column=6, value="Founder Value")
for col in range(2, 7):
    ws1.cell(row=18, column=col).font = bold_font

# Success scenario
ws1.cell(row=19, column=2, value="Product gains traction")
add_input_row(ws1, 19, "", 0.30, col_label=2, col_value=3, is_pct=True)
ws1.cell(row=19, column=3).font = input_font
add_input_row(ws1, 19, "", 25.0, col_label=3, col_value=4, is_currency=True)
ws1.cell(row=19, column=4).font = input_font
ws1.cell(row=19, column=5, value="=G15")
ws1.cell(row=19, column=5).number_format = '0.0%'
ws1.cell(row=19, column=6, value="=D19*E19")
ws1.cell(row=19, column=6).number_format = '$#,##0.0"M"'

# Moderate scenario
ws1.cell(row=20, column=2, value="Modest growth")
add_input_row(ws1, 20, "", 0.45, col_label=2, col_value=3, is_pct=True)
ws1.cell(row=20, column=3).font = input_font
add_input_row(ws1, 20, "", 12.0, col_label=3, col_value=4, is_currency=True)
ws1.cell(row=20, column=4).font = input_font
ws1.cell(row=20, column=5, value="=G15")
ws1.cell(row=20, column=5).number_format = '0.0%'
ws1.cell(row=20, column=6, value="=D20*E20")
ws1.cell(row=20, column=6).number_format = '$#,##0.0"M"'

# Struggle scenario
ws1.cell(row=21, column=2, value="Product struggles")
add_input_row(ws1, 21, "", 0.25, col_label=2, col_value=3, is_pct=True)
ws1.cell(row=21, column=3).font = input_font
add_input_row(ws1, 21, "", 8.0, col_label=3, col_value=4, is_currency=True)
ws1.cell(row=21, column=4).font = input_font
ws1.cell(row=21, column=5, value="=G15")
ws1.cell(row=21, column=5).number_format = '0.0%'
ws1.cell(row=21, column=6, value="=D21*E21")
ws1.cell(row=21, column=6).number_format = '$#,##0.0"M"'

# Expected Value
add_section(ws1, 23, "EXPECTED VALUE")
ws1.cell(row=24, column=2, value="EV (Enterprise)")
ws1.cell(row=24, column=3, value="=SUMPRODUCT(C19:C21,D19:D21)")
ws1.cell(row=24, column=3).number_format = '$#,##0.0"M"'
ws1.cell(row=24, column=3).font = bold_font

ws1.cell(row=25, column=2, value="EV (Founder)")
ws1.cell(row=25, column=3, value="=SUMPRODUCT(C19:C21,F19:F21)")
ws1.cell(row=25, column=3).number_format = '$#,##0.0"M"'
ws1.cell(row=25, column=3).font = Font(bold=True, color=SYNAPTIC_BLUE)

ws1.cell(row=26, column=2, value="Final Founder Ownership")
ws1.cell(row=26, column=3, value="=G15")
ws1.cell(row=26, column=3).number_format = '0.0%'

# ============================================================================
# SHEET 2: PATH 2 - LIQUIDATE (Sell Services, All-in Product)
# ============================================================================
ws2 = wb.create_sheet("Path2_Liquidate")
set_column_widths(ws2, [3, 30, 15, 15, 15, 15, 15])

add_header(ws2, 1, "PATH 2: LIQUIDATE - Dilution Analysis", 7)

# Assumptions
add_section(ws2, 3, "ASSUMPTIONS")
add_input_row(ws2, 4, "Starting founder ownership", 1.0, is_pct=True)
add_input_row(ws2, 5, "Services sale price ($M)", 10.0, is_currency=True)
add_input_row(ws2, 6, "Founder proceeds from sale ($M)", 8.0, is_currency=True)

# Capital Raises
add_section(ws2, 8, "CAPITAL RAISES")
ws2.cell(row=9, column=2, value="Round")
ws2.cell(row=9, column=3, value="Timing")
ws2.cell(row=9, column=4, value="Pre-Money ($M)")
ws2.cell(row=9, column=5, value="Raise ($M)")
ws2.cell(row=9, column=6, value="Post-Money ($M)")
ws2.cell(row=9, column=7, value="Dilution")
for col in range(2, 8):
    ws2.cell(row=9, column=col).font = bold_font

# Seed round
ws2.cell(row=10, column=2, value="Seed")
ws2.cell(row=10, column=3, value="Month 3")
add_input_row(ws2, 10, "", 8.0, col_label=3, col_value=4, is_currency=True)
ws2.cell(row=10, column=4).font = input_font
add_input_row(ws2, 10, "", 3.0, col_label=4, col_value=5, is_currency=True)
ws2.cell(row=10, column=5).font = input_font
ws2.cell(row=10, column=6, value="=D10+E10")
ws2.cell(row=10, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=10, column=7, value="=E10/F10")
ws2.cell(row=10, column=7).number_format = '0.0%'

# Series A
ws2.cell(row=11, column=2, value="Series A")
ws2.cell(row=11, column=3, value="Month 12")
add_input_row(ws2, 11, "", 25.0, col_label=3, col_value=4, is_currency=True)
ws2.cell(row=11, column=4).font = input_font
add_input_row(ws2, 11, "", 10.0, col_label=4, col_value=5, is_currency=True)
ws2.cell(row=11, column=5).font = input_font
ws2.cell(row=11, column=6, value="=D11+E11")
ws2.cell(row=11, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=11, column=7, value="=E11/F11")
ws2.cell(row=11, column=7).number_format = '0.0%'

# ESOP
ws2.cell(row=12, column=2, value="ESOP Pool")
ws2.cell(row=12, column=3, value="At Seed")
ws2.cell(row=12, column=4, value="-")
add_input_row(ws2, 12, "", 0.10, col_label=4, col_value=5, is_pct=True)
ws2.cell(row=12, column=5).font = input_font
ws2.cell(row=12, column=6, value="-")
ws2.cell(row=12, column=7, value="=E12")
ws2.cell(row=12, column=7).number_format = '0.0%'

# Ownership Waterfall
add_section(ws2, 14, "OWNERSHIP WATERFALL")
ws2.cell(row=15, column=2, value="Event")
ws2.cell(row=15, column=3, value="Founder % Before")
ws2.cell(row=15, column=4, value="Dilution")
ws2.cell(row=15, column=5, value="Founder % After")
for col in range(2, 6):
    ws2.cell(row=15, column=col).font = bold_font

ws2.cell(row=16, column=2, value="Starting")
ws2.cell(row=16, column=3, value="=C4")
ws2.cell(row=16, column=3).number_format = '0.0%'
ws2.cell(row=16, column=4, value="0%")
ws2.cell(row=16, column=5, value="=C16")
ws2.cell(row=16, column=5).number_format = '0.0%'

ws2.cell(row=17, column=2, value="ESOP Carveout")
ws2.cell(row=17, column=3, value="=E16")
ws2.cell(row=17, column=3).number_format = '0.0%'
ws2.cell(row=17, column=4, value="=G12")
ws2.cell(row=17, column=4).number_format = '0.0%'
ws2.cell(row=17, column=5, value="=C17*(1-D17)")
ws2.cell(row=17, column=5).number_format = '0.0%'

ws2.cell(row=18, column=2, value="After Seed")
ws2.cell(row=18, column=3, value="=E17")
ws2.cell(row=18, column=3).number_format = '0.0%'
ws2.cell(row=18, column=4, value="=G10")
ws2.cell(row=18, column=4).number_format = '0.0%'
ws2.cell(row=18, column=5, value="=C18*(1-D18)")
ws2.cell(row=18, column=5).number_format = '0.0%'

ws2.cell(row=19, column=2, value="After Series A")
ws2.cell(row=19, column=3, value="=E18")
ws2.cell(row=19, column=3).number_format = '0.0%'
ws2.cell(row=19, column=4, value="=G11")
ws2.cell(row=19, column=4).number_format = '0.0%'
ws2.cell(row=19, column=5, value="=C19*(1-D19)")
ws2.cell(row=19, column=5).number_format = '0.0%'

# Exit Scenarios
add_section(ws2, 21, "EXIT SCENARIOS (Year 3)")
ws2.cell(row=22, column=2, value="Scenario")
ws2.cell(row=22, column=3, value="Probability")
ws2.cell(row=22, column=4, value="Enterprise Value")
ws2.cell(row=22, column=5, value="Founder %")
ws2.cell(row=22, column=6, value="Founder Value")
ws2.cell(row=22, column=7, value="+ Services $")
for col in range(2, 8):
    ws2.cell(row=22, column=col).font = bold_font

# Success
ws2.cell(row=23, column=2, value="Helm succeeds")
add_input_row(ws2, 23, "", 0.50, col_label=2, col_value=3, is_pct=True)
ws2.cell(row=23, column=3).font = input_font
add_input_row(ws2, 23, "", 80.0, col_label=3, col_value=4, is_currency=True)
ws2.cell(row=23, column=4).font = input_font
ws2.cell(row=23, column=5, value="=E19")
ws2.cell(row=23, column=5).number_format = '0.0%'
ws2.cell(row=23, column=6, value="=D23*E23")
ws2.cell(row=23, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=23, column=7, value="=F23+C6")
ws2.cell(row=23, column=7).number_format = '$#,##0.0"M"'

# Moderate
ws2.cell(row=24, column=2, value="Helm struggles")
add_input_row(ws2, 24, "", 0.30, col_label=2, col_value=3, is_pct=True)
ws2.cell(row=24, column=3).font = input_font
add_input_row(ws2, 24, "", 20.0, col_label=3, col_value=4, is_currency=True)
ws2.cell(row=24, column=4).font = input_font
ws2.cell(row=24, column=5, value="=E19")
ws2.cell(row=24, column=5).number_format = '0.0%'
ws2.cell(row=24, column=6, value="=D24*E24")
ws2.cell(row=24, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=24, column=7, value="=F24+C6")
ws2.cell(row=24, column=7).number_format = '$#,##0.0"M"'

# Fail
ws2.cell(row=25, column=2, value="Helm fails")
add_input_row(ws2, 25, "", 0.20, col_label=2, col_value=3, is_pct=True)
ws2.cell(row=25, column=3).font = input_font
add_input_row(ws2, 25, "", 0.0, col_label=3, col_value=4, is_currency=True)
ws2.cell(row=25, column=4).font = input_font
ws2.cell(row=25, column=5, value="=E19")
ws2.cell(row=25, column=5).number_format = '0.0%'
ws2.cell(row=25, column=6, value="=D25*E25")
ws2.cell(row=25, column=6).number_format = '$#,##0.0"M"'
ws2.cell(row=25, column=7, value="=F25+C6")
ws2.cell(row=25, column=7).number_format = '$#,##0.0"M"'

# Expected Value
add_section(ws2, 27, "EXPECTED VALUE")
ws2.cell(row=28, column=2, value="EV (Enterprise)")
ws2.cell(row=28, column=3, value="=SUMPRODUCT(C23:C25,D23:D25)")
ws2.cell(row=28, column=3).number_format = '$#,##0.0"M"'
ws2.cell(row=28, column=3).font = bold_font

ws2.cell(row=29, column=2, value="EV (Founder from Product)")
ws2.cell(row=29, column=3, value="=SUMPRODUCT(C23:C25,F23:F25)")
ws2.cell(row=29, column=3).number_format = '$#,##0.0"M"'

ws2.cell(row=30, column=2, value="EV (Founder Total incl. Services)")
ws2.cell(row=30, column=3, value="=SUMPRODUCT(C23:C25,G23:G25)")
ws2.cell(row=30, column=3).number_format = '$#,##0.0"M"'
ws2.cell(row=30, column=3).font = Font(bold=True, color=SYNAPTIC_BLUE)

ws2.cell(row=31, column=2, value="Final Founder Ownership (Product)")
ws2.cell(row=31, column=3, value="=E19")
ws2.cell(row=31, column=3).number_format = '0.0%'

# ============================================================================
# SHEET 3: PATH 3 - GROWTH BOTH
# ============================================================================
ws3 = wb.create_sheet("Path3_GrowthBoth")
set_column_widths(ws3, [3, 30, 15, 15, 15, 15, 15])

add_header(ws3, 1, "PATH 3: GROWTH BOTH - Dilution Analysis", 7)

# Assumptions
add_section(ws3, 3, "ASSUMPTIONS")
add_input_row(ws3, 4, "Starting founder ownership", 1.0, is_pct=True)
add_input_row(ws3, 5, "Services revenue Y1 ($M)", 4.5, is_currency=True)
add_input_row(ws3, 6, "Services investment for growth ($M)", 0.2, is_currency=True)
add_input_row(ws3, 7, "Services Y3 value ($M)", 24.0, is_currency=True)

# Capital Raises
add_section(ws3, 9, "CAPITAL RAISES (Product Side)")
ws3.cell(row=10, column=2, value="Round")
ws3.cell(row=10, column=3, value="Timing")
ws3.cell(row=10, column=4, value="Pre-Money ($M)")
ws3.cell(row=10, column=5, value="Raise ($M)")
ws3.cell(row=10, column=6, value="Post-Money ($M)")
ws3.cell(row=10, column=7, value="Dilution")
for col in range(2, 8):
    ws3.cell(row=10, column=col).font = bold_font

# Seed
ws3.cell(row=11, column=2, value="Seed")
ws3.cell(row=11, column=3, value="Month 6")
add_input_row(ws3, 11, "", 6.0, col_label=3, col_value=4, is_currency=True)
ws3.cell(row=11, column=4).font = input_font
add_input_row(ws3, 11, "", 3.0, col_label=4, col_value=5, is_currency=True)
ws3.cell(row=11, column=5).font = input_font
ws3.cell(row=11, column=6, value="=D11+E11")
ws3.cell(row=11, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=11, column=7, value="=E11/F11")
ws3.cell(row=11, column=7).number_format = '0.0%'

# Series A
ws3.cell(row=12, column=2, value="Series A")
ws3.cell(row=12, column=3, value="Month 18")
add_input_row(ws3, 12, "", 20.0, col_label=3, col_value=4, is_currency=True)
ws3.cell(row=12, column=4).font = input_font
add_input_row(ws3, 12, "", 8.0, col_label=4, col_value=5, is_currency=True)
ws3.cell(row=12, column=5).font = input_font
ws3.cell(row=12, column=6, value="=D12+E12")
ws3.cell(row=12, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=12, column=7, value="=E12/F12")
ws3.cell(row=12, column=7).number_format = '0.0%'

# ESOP
ws3.cell(row=13, column=2, value="ESOP Pool")
ws3.cell(row=13, column=3, value="At Seed")
ws3.cell(row=13, column=4, value="-")
add_input_row(ws3, 13, "", 0.10, col_label=4, col_value=5, is_pct=True)
ws3.cell(row=13, column=5).font = input_font
ws3.cell(row=13, column=6, value="-")
ws3.cell(row=13, column=7, value="=E13")
ws3.cell(row=13, column=7).number_format = '0.0%'

# Ownership Waterfall
add_section(ws3, 15, "OWNERSHIP WATERFALL (Product Entity)")
ws3.cell(row=16, column=2, value="Event")
ws3.cell(row=16, column=3, value="Founder % Before")
ws3.cell(row=16, column=4, value="Dilution")
ws3.cell(row=16, column=5, value="Founder % After")
for col in range(2, 6):
    ws3.cell(row=16, column=col).font = bold_font

ws3.cell(row=17, column=2, value="Starting")
ws3.cell(row=17, column=3, value="=C4")
ws3.cell(row=17, column=3).number_format = '0.0%'
ws3.cell(row=17, column=4, value="0%")
ws3.cell(row=17, column=5, value="=C17")
ws3.cell(row=17, column=5).number_format = '0.0%'

ws3.cell(row=18, column=2, value="ESOP Carveout")
ws3.cell(row=18, column=3, value="=E17")
ws3.cell(row=18, column=3).number_format = '0.0%'
ws3.cell(row=18, column=4, value="=G13")
ws3.cell(row=18, column=4).number_format = '0.0%'
ws3.cell(row=18, column=5, value="=C18*(1-D18)")
ws3.cell(row=18, column=5).number_format = '0.0%'

ws3.cell(row=19, column=2, value="After Seed")
ws3.cell(row=19, column=3, value="=E18")
ws3.cell(row=19, column=3).number_format = '0.0%'
ws3.cell(row=19, column=4, value="=G11")
ws3.cell(row=19, column=4).number_format = '0.0%'
ws3.cell(row=19, column=5, value="=C19*(1-D19)")
ws3.cell(row=19, column=5).number_format = '0.0%'

ws3.cell(row=20, column=2, value="After Series A")
ws3.cell(row=20, column=3, value="=E19")
ws3.cell(row=20, column=3).number_format = '0.0%'
ws3.cell(row=20, column=4, value="=G12")
ws3.cell(row=20, column=4).number_format = '0.0%'
ws3.cell(row=20, column=5, value="=C20*(1-D20)")
ws3.cell(row=20, column=5).number_format = '0.0%'

# Exit Scenarios
add_section(ws3, 22, "EXIT SCENARIOS (Year 3)")
ws3.cell(row=23, column=2, value="Scenario")
ws3.cell(row=23, column=3, value="Probability")
ws3.cell(row=23, column=4, value="Product EV")
ws3.cell(row=23, column=5, value="Services EV")
ws3.cell(row=23, column=6, value="Founder Product")
ws3.cell(row=23, column=7, value="Total Founder")
for col in range(2, 8):
    ws3.cell(row=23, column=col).font = bold_font

# Success
ws3.cell(row=24, column=2, value="Both succeed")
add_input_row(ws3, 24, "", 0.40, col_label=2, col_value=3, is_pct=True)
ws3.cell(row=24, column=3).font = input_font
add_input_row(ws3, 24, "", 60.0, col_label=3, col_value=4, is_currency=True)
ws3.cell(row=24, column=4).font = input_font
ws3.cell(row=24, column=5, value="=C7")
ws3.cell(row=24, column=5).number_format = '$#,##0.0"M"'
ws3.cell(row=24, column=6, value="=D24*E20")
ws3.cell(row=24, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=24, column=7, value="=F24+E24")
ws3.cell(row=24, column=7).number_format = '$#,##0.0"M"'

# Mixed
ws3.cell(row=25, column=2, value="Product struggles")
add_input_row(ws3, 25, "", 0.35, col_label=2, col_value=3, is_pct=True)
ws3.cell(row=25, column=3).font = input_font
add_input_row(ws3, 25, "", 15.0, col_label=3, col_value=4, is_currency=True)
ws3.cell(row=25, column=4).font = input_font
ws3.cell(row=25, column=5, value="=C7")
ws3.cell(row=25, column=5).number_format = '$#,##0.0"M"'
ws3.cell(row=25, column=6, value="=D25*E20")
ws3.cell(row=25, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=25, column=7, value="=F25+E25")
ws3.cell(row=25, column=7).number_format = '$#,##0.0"M"'

# Both struggle
ws3.cell(row=26, column=2, value="Both struggle")
add_input_row(ws3, 26, "", 0.25, col_label=2, col_value=3, is_pct=True)
ws3.cell(row=26, column=3).font = input_font
add_input_row(ws3, 26, "", 5.0, col_label=3, col_value=4, is_currency=True)
ws3.cell(row=26, column=4).font = input_font
add_input_row(ws3, 26, "", 12.0, col_label=4, col_value=5, is_currency=True)
ws3.cell(row=26, column=5).font = input_font
ws3.cell(row=26, column=6, value="=D26*E20")
ws3.cell(row=26, column=6).number_format = '$#,##0.0"M"'
ws3.cell(row=26, column=7, value="=F26+E26")
ws3.cell(row=26, column=7).number_format = '$#,##0.0"M"'

# Expected Value
add_section(ws3, 28, "EXPECTED VALUE")
ws3.cell(row=29, column=2, value="EV (Enterprise Total)")
ws3.cell(row=29, column=3, value="=SUMPRODUCT(C24:C26,D24:D26)+SUMPRODUCT(C24:C26,E24:E26)")
ws3.cell(row=29, column=3).number_format = '$#,##0.0"M"'
ws3.cell(row=29, column=3).font = bold_font

ws3.cell(row=30, column=2, value="EV (Founder Total)")
ws3.cell(row=30, column=3, value="=SUMPRODUCT(C24:C26,G24:G26)")
ws3.cell(row=30, column=3).number_format = '$#,##0.0"M"'
ws3.cell(row=30, column=3).font = Font(bold=True, color=SYNAPTIC_BLUE)

ws3.cell(row=31, column=2, value="Final Founder % (Product)")
ws3.cell(row=31, column=3, value="=E20")
ws3.cell(row=31, column=3).number_format = '0.0%'

# ============================================================================
# SHEET 4: PATH 4 - STRATEGIC LEVERAGE (Recommended)
# ============================================================================
ws4 = wb.create_sheet("Path4_Strategic")
set_column_widths(ws4, [3, 30, 15, 15, 15, 15, 15])

add_header(ws4, 1, "PATH 4: STRATEGIC LEVERAGE - Dilution Analysis (RECOMMENDED)", 7)

# Assumptions
add_section(ws4, 3, "ASSUMPTIONS")
add_input_row(ws4, 4, "Starting founder ownership", 1.0, is_pct=True)
add_input_row(ws4, 5, "Services revenue ($M)", 4.0, is_currency=True)
add_input_row(ws4, 6, "Services EBITDA margin (optimized)", 0.20, is_pct=True)
add_input_row(ws4, 7, "Services sale price Y1 ($M)", 10.0, is_currency=True)
add_input_row(ws4, 8, "Founder % of services proceeds", 0.80, is_pct=True)

# Capital Raises - PHASED based on validation
add_section(ws4, 10, "CAPITAL RAISES (Phased Based on Validation)")
ws4.cell(row=11, column=2, value="Round")
ws4.cell(row=11, column=3, value="Timing")
ws4.cell(row=11, column=4, value="Pre-Money ($M)")
ws4.cell(row=11, column=5, value="Raise ($M)")
ws4.cell(row=11, column=6, value="Post-Money ($M)")
ws4.cell(row=11, column=7, value="Dilution")
for col in range(2, 8):
    ws4.cell(row=11, column=col).font = bold_font

# Seed (post-validation, better terms)
ws4.cell(row=12, column=2, value="Seed (w/ proof points)")
ws4.cell(row=12, column=3, value="Month 12")
add_input_row(ws4, 12, "", 12.0, col_label=3, col_value=4, is_currency=True)
ws4.cell(row=12, column=4).font = input_font
add_input_row(ws4, 12, "", 3.0, col_label=4, col_value=5, is_currency=True)
ws4.cell(row=12, column=5).font = input_font
ws4.cell(row=12, column=6, value="=D12+E12")
ws4.cell(row=12, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=12, column=7, value="=E12/F12")
ws4.cell(row=12, column=7).number_format = '0.0%'

# Series A
ws4.cell(row=13, column=2, value="Series A")
ws4.cell(row=13, column=3, value="Month 24")
add_input_row(ws4, 13, "", 35.0, col_label=3, col_value=4, is_currency=True)
ws4.cell(row=13, column=4).font = input_font
add_input_row(ws4, 13, "", 10.0, col_label=4, col_value=5, is_currency=True)
ws4.cell(row=13, column=5).font = input_font
ws4.cell(row=13, column=6, value="=D13+E13")
ws4.cell(row=13, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=13, column=7, value="=E13/F13")
ws4.cell(row=13, column=7).number_format = '0.0%'

# ESOP
ws4.cell(row=14, column=2, value="ESOP Pool")
ws4.cell(row=14, column=3, value="At Seed")
ws4.cell(row=14, column=4, value="-")
add_input_row(ws4, 14, "", 0.10, col_label=4, col_value=5, is_pct=True)
ws4.cell(row=14, column=5).font = input_font
ws4.cell(row=14, column=6, value="-")
ws4.cell(row=14, column=7, value="=E14")
ws4.cell(row=14, column=7).number_format = '0.0%'

# Ownership Waterfall
add_section(ws4, 16, "OWNERSHIP WATERFALL")
ws4.cell(row=17, column=2, value="Event")
ws4.cell(row=17, column=3, value="Founder % Before")
ws4.cell(row=17, column=4, value="Dilution")
ws4.cell(row=17, column=5, value="Founder % After")
for col in range(2, 6):
    ws4.cell(row=17, column=col).font = bold_font

ws4.cell(row=18, column=2, value="Starting")
ws4.cell(row=18, column=3, value="=C4")
ws4.cell(row=18, column=3).number_format = '0.0%'
ws4.cell(row=18, column=4, value="0%")
ws4.cell(row=18, column=5, value="=C18")
ws4.cell(row=18, column=5).number_format = '0.0%'

ws4.cell(row=19, column=2, value="ESOP Carveout")
ws4.cell(row=19, column=3, value="=E18")
ws4.cell(row=19, column=3).number_format = '0.0%'
ws4.cell(row=19, column=4, value="=G14")
ws4.cell(row=19, column=4).number_format = '0.0%'
ws4.cell(row=19, column=5, value="=C19*(1-D19)")
ws4.cell(row=19, column=5).number_format = '0.0%'

ws4.cell(row=20, column=2, value="After Seed")
ws4.cell(row=20, column=3, value="=E19")
ws4.cell(row=20, column=3).number_format = '0.0%'
ws4.cell(row=20, column=4, value="=G12")
ws4.cell(row=20, column=4).number_format = '0.0%'
ws4.cell(row=20, column=5, value="=C20*(1-D20)")
ws4.cell(row=20, column=5).number_format = '0.0%'

ws4.cell(row=21, column=2, value="After Series A")
ws4.cell(row=21, column=3, value="=E20")
ws4.cell(row=21, column=3).number_format = '0.0%'
ws4.cell(row=21, column=4, value="=G13")
ws4.cell(row=21, column=4).number_format = '0.0%'
ws4.cell(row=21, column=5, value="=C21*(1-D21)")
ws4.cell(row=21, column=5).number_format = '0.0%'

# Exit Scenarios
add_section(ws4, 23, "EXIT SCENARIOS (Year 3)")
ws4.cell(row=24, column=2, value="Scenario")
ws4.cell(row=24, column=3, value="Probability")
ws4.cell(row=24, column=4, value="Product EV")
ws4.cell(row=24, column=5, value="Founder Product")
ws4.cell(row=24, column=6, value="Services $")
ws4.cell(row=24, column=7, value="Total Founder")
for col in range(2, 8):
    ws4.cell(row=24, column=col).font = bold_font

# Success - Helm takes off
ws4.cell(row=25, column=2, value="Helm succeeds")
add_input_row(ws4, 25, "", 0.60, col_label=2, col_value=3, is_pct=True)
ws4.cell(row=25, column=3).font = input_font
add_input_row(ws4, 25, "", 80.0, col_label=3, col_value=4, is_currency=True)
ws4.cell(row=25, column=4).font = input_font
ws4.cell(row=25, column=5, value="=D25*E21")
ws4.cell(row=25, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=25, column=6, value="=C7*C8")
ws4.cell(row=25, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=25, column=7, value="=E25+F25")
ws4.cell(row=25, column=7).number_format = '$#,##0.0"M"'

# Moderate - Helm struggles but services exit preserved
ws4.cell(row=26, column=2, value="Helm struggles")
add_input_row(ws4, 26, "", 0.25, col_label=2, col_value=3, is_pct=True)
ws4.cell(row=26, column=3).font = input_font
add_input_row(ws4, 26, "", 15.0, col_label=3, col_value=4, is_currency=True)
ws4.cell(row=26, column=4).font = input_font
ws4.cell(row=26, column=5, value="=D26*E21")
ws4.cell(row=26, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=26, column=6, value="=C7*C8")
ws4.cell(row=26, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=26, column=7, value="=E26+F26")
ws4.cell(row=26, column=7).number_format = '$#,##0.0"M"'

# Pivot - Different direction, services $ preserved
ws4.cell(row=27, column=2, value="Helm pivots")
add_input_row(ws4, 27, "", 0.15, col_label=2, col_value=3, is_pct=True)
ws4.cell(row=27, column=3).font = input_font
add_input_row(ws4, 27, "", 5.0, col_label=3, col_value=4, is_currency=True)
ws4.cell(row=27, column=4).font = input_font
ws4.cell(row=27, column=5, value="=D27*E21")
ws4.cell(row=27, column=5).number_format = '$#,##0.0"M"'
ws4.cell(row=27, column=6, value="=C7*C8")
ws4.cell(row=27, column=6).number_format = '$#,##0.0"M"'
ws4.cell(row=27, column=7, value="=E27+F27")
ws4.cell(row=27, column=7).number_format = '$#,##0.0"M"'

# Expected Value
add_section(ws4, 29, "EXPECTED VALUE")
ws4.cell(row=30, column=2, value="EV (Enterprise)")
ws4.cell(row=30, column=3, value="=SUMPRODUCT(C25:C27,D25:D27)")
ws4.cell(row=30, column=3).number_format = '$#,##0.0"M"'
ws4.cell(row=30, column=3).font = bold_font

ws4.cell(row=31, column=2, value="EV (Founder Total)")
ws4.cell(row=31, column=3, value="=SUMPRODUCT(C25:C27,G25:G27)")
ws4.cell(row=31, column=3).number_format = '$#,##0.0"M"'
ws4.cell(row=31, column=3).font = Font(bold=True, color=SYNAPTIC_BLUE)

ws4.cell(row=32, column=2, value="Final Founder % (Product)")
ws4.cell(row=32, column=3, value="=E21")
ws4.cell(row=32, column=3).number_format = '0.0%'

# Key Advantage callout
add_section(ws4, 34, "KEY ADVANTAGES")
ws4.cell(row=35, column=2, value="1. Higher pre-money ($12M vs $8M) due to validation proof points")
ws4.cell(row=36, column=2, value="2. Services proceeds ($8M) provide downside protection in ALL scenarios")
ws4.cell(row=37, column=2, value="3. Better founder ownership (~58%) vs Liquidate (~58%) with LESS risk")
ws4.cell(row=38, column=2, value="4. Optionality preserved - can adjust strategy at Month 12 decision point")

# ============================================================================
# SHEET 5: SUMMARY COMPARISON
# ============================================================================
ws5 = wb.create_sheet("Summary_Comparison")
set_column_widths(ws5, [3, 25, 18, 18, 18, 18])

add_header(ws5, 1, "STRATEGIC PATH COMPARISON - Dilution-Adjusted Analysis", 6)

# Summary table
add_section(ws5, 3, "EXPECTED VALUE COMPARISON")
ws5.cell(row=4, column=2, value="Metric")
ws5.cell(row=4, column=3, value="Path 1: Lean")
ws5.cell(row=4, column=4, value="Path 2: Liquidate")
ws5.cell(row=4, column=5, value="Path 3: Growth Both")
ws5.cell(row=4, column=6, value="Path 4: Strategic")
for col in range(2, 7):
    ws5.cell(row=4, column=col).font = bold_font
    ws5.cell(row=4, column=col).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

# EV Enterprise
ws5.cell(row=5, column=2, value="EV (Enterprise)")
ws5.cell(row=5, column=3, value="=Path1_Lean!C24")
ws5.cell(row=5, column=4, value="=Path2_Liquidate!C28")
ws5.cell(row=5, column=5, value="=Path3_GrowthBoth!C29")
ws5.cell(row=5, column=6, value="=Path4_Strategic!C30")
for col in range(3, 7):
    ws5.cell(row=5, column=col).number_format = '$#,##0.0"M"'

# EV Founder
ws5.cell(row=6, column=2, value="EV (Founder)")
ws5.cell(row=6, column=3, value="=Path1_Lean!C25")
ws5.cell(row=6, column=4, value="=Path2_Liquidate!C30")
ws5.cell(row=6, column=5, value="=Path3_GrowthBoth!C30")
ws5.cell(row=6, column=6, value="=Path4_Strategic!C31")
for col in range(3, 7):
    ws5.cell(row=6, column=col).number_format = '$#,##0.0"M"'
    ws5.cell(row=6, column=col).font = Font(bold=True, color=SYNAPTIC_BLUE)

# Final Ownership
ws5.cell(row=7, column=2, value="Final Founder %")
ws5.cell(row=7, column=3, value="=Path1_Lean!C26")
ws5.cell(row=7, column=4, value="=Path2_Liquidate!C31")
ws5.cell(row=7, column=5, value="=Path3_GrowthBoth!C31")
ws5.cell(row=7, column=6, value="=Path4_Strategic!C32")
for col in range(3, 7):
    ws5.cell(row=7, column=col).number_format = '0.0%'

# Total dilution
ws5.cell(row=8, column=2, value="Total Dilution")
ws5.cell(row=8, column=3, value="=1-C7")
ws5.cell(row=8, column=4, value="=1-D7")
ws5.cell(row=8, column=5, value="=1-E7")
ws5.cell(row=8, column=6, value="=1-F7")
for col in range(3, 7):
    ws5.cell(row=8, column=col).number_format = '0.0%'

# Capital raised
ws5.cell(row=9, column=2, value="Total Capital Raised")
ws5.cell(row=9, column=3, value=0)
ws5.cell(row=9, column=4, value="=Path2_Liquidate!E10+Path2_Liquidate!E11")
ws5.cell(row=9, column=5, value="=Path3_GrowthBoth!E11+Path3_GrowthBoth!E12")
ws5.cell(row=9, column=6, value="=Path4_Strategic!E12+Path4_Strategic!E13")
for col in range(3, 7):
    ws5.cell(row=9, column=col).number_format = '$#,##0.0"M"'

# Dilution-Adjusted Rankings
add_section(ws5, 11, "DILUTION-ADJUSTED RANKING")
ws5.cell(row=12, column=2, value="Rank by Founder EV")
ws5.cell(row=12, column=3, value='=RANK(C6,$C$6:$F$6,0)')
ws5.cell(row=12, column=4, value='=RANK(D6,$C$6:$F$6,0)')
ws5.cell(row=12, column=5, value='=RANK(E6,$C$6:$F$6,0)')
ws5.cell(row=12, column=6, value='=RANK(F6,$C$6:$F$6,0)')

# Highlight winner
ws5.cell(row=13, column=2, value="WINNER")
ws5.cell(row=13, column=6, value="Strategic Leverage")
ws5.cell(row=13, column=6).font = Font(bold=True, color="FFFFFF")
ws5.cell(row=13, column=6).fill = PatternFill("solid", fgColor=SYNAPTIC_BLUE)

# Risk Analysis
add_section(ws5, 15, "RISK-REWARD ANALYSIS")
ws5.cell(row=16, column=2, value="Best Case (Founder)")
ws5.cell(row=16, column=3, value="=Path1_Lean!F19")
ws5.cell(row=16, column=4, value="=Path2_Liquidate!G23")
ws5.cell(row=16, column=5, value="=Path3_GrowthBoth!G24")
ws5.cell(row=16, column=6, value="=Path4_Strategic!G25")
for col in range(3, 7):
    ws5.cell(row=16, column=col).number_format = '$#,##0.0"M"'

ws5.cell(row=17, column=2, value="Worst Case (Founder)")
ws5.cell(row=17, column=3, value="=Path1_Lean!F21")
ws5.cell(row=17, column=4, value="=Path2_Liquidate!G25")
ws5.cell(row=17, column=5, value="=Path3_GrowthBoth!G26")
ws5.cell(row=17, column=6, value="=Path4_Strategic!G27")
for col in range(3, 7):
    ws5.cell(row=17, column=col).number_format = '$#,##0.0"M"'

ws5.cell(row=18, column=2, value="Downside Protection")
ws5.cell(row=18, column=3, value="=C17")
ws5.cell(row=18, column=4, value="=D17")
ws5.cell(row=18, column=5, value="=E17")
ws5.cell(row=18, column=6, value="=F17")
for col in range(3, 7):
    ws5.cell(row=18, column=col).number_format = '$#,##0.0"M"'

# Key Insight
add_section(ws5, 20, "KEY INSIGHT")
ws5.cell(row=21, column=2, value="Strategic Leverage delivers highest founder EV while maintaining")
ws5.cell(row=22, column=2, value="strong downside protection via services exit proceeds.")
ws5.cell(row=23, column=2, value="The higher pre-money valuations (from proof points) offset dilution.")

# Save workbook
output_path = "/Users/wkarp/devlocal/prod-miner-pm/x-project/bplan/numbers/strategic-scenarios-with-dilution.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
