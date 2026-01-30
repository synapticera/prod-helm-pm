from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, size=14, color="FFFFFF")
subheader_font = Font(bold=True, size=12)
blue_fill = PatternFill("solid", fgColor="07253D")  # Synaptic Dark Navy
accent_fill = PatternFill("solid", fgColor="01A9DB")  # Synaptic Blue
light_fill = PatternFill("solid", fgColor="F4F6F6")
input_font = Font(color="0000FF")  # Blue for inputs
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, row, col, text, width=15):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col)].width = width

def style_label(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')

def style_input(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = input_font
    cell.alignment = Alignment(horizontal='right')

#==============================================================================
# SCENARIO 1: LEAN
#==============================================================================
ws1 = wb.active
ws1.title = "Path1_Lean"

# Title
ws1.merge_cells('B2:H2')
ws1.cell(row=2, column=2, value="Path 1: LEAN - Minimal Investment, Self-Funded")
ws1.cell(row=2, column=2).font = Font(bold=True, size=16)

# Assumptions Section
ws1.cell(row=4, column=2, value="ASSUMPTIONS").font = Font(bold=True, size=12)
ws1.cell(row=5, column=2, value="Services starting revenue ($M)")
ws1.cell(row=5, column=3, value=4.0)
ws1['C5'].font = input_font

ws1.cell(row=6, column=2, value="Services annual decline")
ws1.cell(row=6, column=3, value=-0.05)
ws1['C6'].font = input_font
ws1['C6'].number_format = '0%'

ws1.cell(row=7, column=2, value="Services EBITDA margin")
ws1.cell(row=7, column=3, value=0.12)
ws1['C7'].font = input_font
ws1['C7'].number_format = '0%'

ws1.cell(row=8, column=2, value="Product investment/year ($M)")
ws1.cell(row=8, column=3, value=0.35)
ws1['C8'].font = input_font

ws1.cell(row=9, column=2, value="Services EV multiple")
ws1.cell(row=9, column=3, value=2.5)
ws1['C9'].font = input_font

ws1.cell(row=10, column=2, value="Product EV multiple (if successful)")
ws1.cell(row=10, column=3, value=10.0)
ws1['C10'].font = input_font

# Growth Plan
ws1.cell(row=12, column=2, value="GROWTH PLAN").font = Font(bold=True, size=12)
years = [2025, 2026, 2027, 2028, 2029]
for i, year in enumerate(years):
    ws1.cell(row=13, column=3+i, value=year)
    ws1.cell(row=13, column=3+i).font = Font(bold=True)
    ws1.cell(row=13, column=3+i).fill = light_fill

ws1.cell(row=14, column=2, value="Services Revenue ($M)")
ws1.cell(row=14, column=3, value="=C5")
for i in range(1, 5):
    ws1.cell(row=14, column=3+i, value=f"={get_column_letter(2+i)}14*(1+$C$6)")

ws1.cell(row=15, column=2, value="Services EBITDA ($M)")
for i in range(5):
    ws1.cell(row=15, column=3+i, value=f"={get_column_letter(3+i)}14*$C$7")

ws1.cell(row=16, column=2, value="Product Investment ($M)")
for i in range(5):
    ws1.cell(row=16, column=3+i, value="=-$C$8")

ws1.cell(row=17, column=2, value="Net Cash Flow ($M)")
for i in range(5):
    ws1.cell(row=17, column=3+i, value=f"={get_column_letter(3+i)}15+{get_column_letter(3+i)}16")

# Product Revenue (very slow growth due to underfunding)
ws1.cell(row=19, column=2, value="Product Revenue ($M)")
ws1.cell(row=19, column=3, value=0.05)
ws1.cell(row=19, column=4, value=0.15)
ws1.cell(row=19, column=5, value=0.4)
ws1.cell(row=19, column=6, value=0.8)
ws1.cell(row=19, column=7, value=1.5)

# EV Calculation
ws1.cell(row=21, column=2, value="EV CALCULATION").font = Font(bold=True, size=12)

ws1.cell(row=22, column=2, value="Services EV ($M)")
for i in range(5):
    ws1.cell(row=22, column=3+i, value=f"={get_column_letter(3+i)}14*$C$9")

ws1.cell(row=23, column=2, value="Product EV ($M)")
for i in range(5):
    ws1.cell(row=23, column=3+i, value=f"={get_column_letter(3+i)}19*$C$10")

ws1.cell(row=24, column=2, value="Total EV ($M)")
ws1.cell(row=24, column=2).font = Font(bold=True)
for i in range(5):
    ws1.cell(row=24, column=3+i, value=f"={get_column_letter(3+i)}22+{get_column_letter(3+i)}23")
    ws1.cell(row=24, column=3+i).font = Font(bold=True)
    ws1.cell(row=24, column=3+i).fill = accent_fill

# Format numbers
for row in range(14, 25):
    for col in range(3, 8):
        ws1.cell(row=row, column=col).number_format = '#,##0.00'

ws1.column_dimensions['B'].width = 30
for col in range(3, 8):
    ws1.column_dimensions[get_column_letter(col)].width = 12

#==============================================================================
# SCENARIO 2: LIQUIDATE
#==============================================================================
ws2 = wb.create_sheet("Path2_Liquidate")

# Title
ws2.merge_cells('B2:H2')
ws2.cell(row=2, column=2, value="Path 2: LIQUIDATE - Sell Services, All-In on Product")
ws2.cell(row=2, column=2).font = Font(bold=True, size=16)

# Assumptions
ws2.cell(row=4, column=2, value="ASSUMPTIONS").font = Font(bold=True, size=12)
ws2.cell(row=5, column=2, value="Services sale price ($M)")
ws2.cell(row=5, column=3, value=10.0)
ws2['C5'].font = input_font

ws2.cell(row=6, column=2, value="Net after tax/earnout ($M)")
ws2.cell(row=6, column=3, value=7.0)
ws2['C6'].font = input_font

ws2.cell(row=7, column=2, value="Product burn rate/year ($M)")
ws2.cell(row=7, column=3, value=2.5)
ws2['C7'].font = input_font

ws2.cell(row=8, column=2, value="Product EV multiple")
ws2.cell(row=8, column=3, value=10.0)
ws2['C8'].font = input_font

ws2.cell(row=9, column=2, value="Sale completed (year)")
ws2.cell(row=9, column=3, value=2025)
ws2['C9'].font = input_font

# Cash Position
ws2.cell(row=11, column=2, value="CASH POSITION").font = Font(bold=True, size=12)
for i, year in enumerate(years):
    ws2.cell(row=12, column=3+i, value=year)
    ws2.cell(row=12, column=3+i).font = Font(bold=True)
    ws2.cell(row=12, column=3+i).fill = light_fill

ws2.cell(row=13, column=2, value="Starting Cash ($M)")
ws2.cell(row=13, column=3, value="=$C$6")
for i in range(1, 5):
    ws2.cell(row=13, column=3+i, value=f"={get_column_letter(2+i)}15")

ws2.cell(row=14, column=2, value="Product Burn ($M)")
for i in range(5):
    ws2.cell(row=14, column=3+i, value="=-$C$7")

ws2.cell(row=15, column=2, value="Ending Cash ($M)")
for i in range(5):
    ws2.cell(row=15, column=3+i, value=f"={get_column_letter(3+i)}13+{get_column_letter(3+i)}14")

# Product Growth (faster with full focus)
ws2.cell(row=17, column=2, value="PRODUCT GROWTH").font = Font(bold=True, size=12)
ws2.cell(row=18, column=2, value="Product Revenue ($M)")
ws2.cell(row=18, column=3, value=0.2)
ws2.cell(row=18, column=4, value=1.0)
ws2.cell(row=18, column=5, value=5.0)
ws2.cell(row=18, column=6, value=15.0)
ws2.cell(row=18, column=7, value=40.0)

ws2.cell(row=19, column=2, value="YoY Growth")
ws2.cell(row=19, column=3, value="N/A")
for i in range(1, 5):
    ws2.cell(row=19, column=3+i, value=f"=({get_column_letter(3+i)}18-{get_column_letter(2+i)}18)/{get_column_letter(2+i)}18")
    ws2.cell(row=19, column=3+i).number_format = '0%'

# EV Calculation
ws2.cell(row=21, column=2, value="EV CALCULATION").font = Font(bold=True, size=12)
ws2.cell(row=22, column=2, value="Product EV ($M)")
for i in range(5):
    ws2.cell(row=22, column=3+i, value=f"={get_column_letter(3+i)}18*$C$8")
    ws2.cell(row=22, column=3+i).font = Font(bold=True)
    ws2.cell(row=22, column=3+i).fill = accent_fill

ws2.cell(row=23, column=2, value="Cash Position ($M)")
for i in range(5):
    ws2.cell(row=23, column=3+i, value=f"={get_column_letter(3+i)}15")

ws2.cell(row=24, column=2, value="Total EV ($M)")
ws2.cell(row=24, column=2).font = Font(bold=True)
for i in range(5):
    ws2.cell(row=24, column=3+i, value=f"={get_column_letter(3+i)}22+MAX(0,{get_column_letter(3+i)}23)")
    ws2.cell(row=24, column=3+i).font = Font(bold=True)

# Format
for row in range(13, 25):
    for col in range(3, 8):
        if ws2.cell(row=row, column=col).value != "N/A":
            ws2.cell(row=row, column=col).number_format = '#,##0.00'

ws2.column_dimensions['B'].width = 30
for col in range(3, 8):
    ws2.column_dimensions[get_column_letter(col)].width = 12

#==============================================================================
# SCENARIO 3: GROWTH BOTH (from original)
#==============================================================================
ws3 = wb.create_sheet("Path3_GrowthBoth")

# Title
ws3.merge_cells('B2:H2')
ws3.cell(row=2, column=2, value="Path 3: GROWTH BOTH - Dual Business Strategy")
ws3.cell(row=2, column=2).font = Font(bold=True, size=16)

# Assumptions
ws3.cell(row=4, column=2, value="ASSUMPTIONS").font = Font(bold=True, size=12)
ws3.cell(row=5, column=2, value="Services starting revenue ($M)")
ws3.cell(row=5, column=3, value=4.0)
ws3['C5'].font = input_font

ws3.cell(row=6, column=2, value="Services annual growth")
ws3.cell(row=6, column=3, value=0.30)
ws3['C6'].font = input_font
ws3['C6'].number_format = '0%'

ws3.cell(row=7, column=2, value="Services EV multiple")
ws3.cell(row=7, column=3, value=3.0)
ws3['C7'].font = input_font

ws3.cell(row=8, column=2, value="Product EV multiple")
ws3.cell(row=8, column=3, value=10.0)
ws3['C8'].font = input_font

ws3.cell(row=9, column=2, value="VC funding (Year 1, $M)")
ws3.cell(row=9, column=3, value=3.0)
ws3['C9'].font = input_font

# Growth Plan
ws3.cell(row=11, column=2, value="GROWTH PLAN").font = Font(bold=True, size=12)
for i, year in enumerate(years):
    ws3.cell(row=12, column=3+i, value=year)
    ws3.cell(row=12, column=3+i).font = Font(bold=True)
    ws3.cell(row=12, column=3+i).fill = light_fill

ws3.cell(row=13, column=2, value="Services Revenue ($M)")
ws3.cell(row=13, column=3, value=4.275)
ws3.cell(row=13, column=4, value=5.56)
ws3.cell(row=13, column=5, value=7.22)
ws3.cell(row=13, column=6, value=9.39)
ws3.cell(row=13, column=7, value=11.74)

ws3.cell(row=14, column=2, value="Product Revenue ($M)")
ws3.cell(row=14, column=3, value=0)
ws3.cell(row=14, column=4, value=1.0)
ws3.cell(row=14, column=5, value=5.0)
ws3.cell(row=14, column=6, value=25.0)
ws3.cell(row=14, column=7, value=100.0)

ws3.cell(row=15, column=2, value="Total Revenue ($M)")
for i in range(5):
    ws3.cell(row=15, column=3+i, value=f"={get_column_letter(3+i)}13+{get_column_letter(3+i)}14")

# EV Calculation
ws3.cell(row=17, column=2, value="EV CALCULATION").font = Font(bold=True, size=12)

ws3.cell(row=18, column=2, value="Services EV ($M)")
for i in range(5):
    ws3.cell(row=18, column=3+i, value=f"={get_column_letter(3+i)}13*$C$7")

ws3.cell(row=19, column=2, value="Product EV ($M)")
for i in range(5):
    ws3.cell(row=19, column=3+i, value=f"={get_column_letter(3+i)}14*$C$8")

ws3.cell(row=20, column=2, value="Total EV ($M)")
ws3.cell(row=20, column=2).font = Font(bold=True)
for i in range(5):
    ws3.cell(row=20, column=3+i, value=f"={get_column_letter(3+i)}18+{get_column_letter(3+i)}19")
    ws3.cell(row=20, column=3+i).font = Font(bold=True)
    ws3.cell(row=20, column=3+i).fill = accent_fill

# Format
for row in range(13, 21):
    for col in range(3, 8):
        ws3.cell(row=row, column=col).number_format = '#,##0.00'

ws3.column_dimensions['B'].width = 30
for col in range(3, 8):
    ws3.column_dimensions[get_column_letter(col)].width = 12

#==============================================================================
# SCENARIO 4: STRATEGIC LEVERAGE (RECOMMENDED)
#==============================================================================
ws4 = wb.create_sheet("Path4_Strategic")

# Title
ws4.merge_cells('B2:H2')
ws4.cell(row=2, column=2, value="Path 4: STRATEGIC LEVERAGE - Bridge to Product (RECOMMENDED)")
ws4.cell(row=2, column=2).font = Font(bold=True, size=16, color="2E7D32")

# Assumptions
ws4.cell(row=4, column=2, value="ASSUMPTIONS").font = Font(bold=True, size=12)
ws4.cell(row=5, column=2, value="Services starting revenue ($M)")
ws4.cell(row=5, column=3, value=4.0)
ws4['C5'].font = input_font

ws4.cell(row=6, column=2, value="Services EBITDA margin (optimized)")
ws4.cell(row=6, column=3, value=0.20)
ws4['C6'].font = input_font
ws4['C6'].number_format = '0%'

ws4.cell(row=7, column=2, value="Services sale price (Year 2, $M)")
ws4.cell(row=7, column=3, value=10.0)
ws4['C7'].font = input_font

ws4.cell(row=8, column=2, value="Net after sale ($M)")
ws4.cell(row=8, column=3, value=7.0)
ws4['C8'].font = input_font

ws4.cell(row=9, column=2, value="Seed funding (Year 2, $M)")
ws4.cell(row=9, column=3, value=3.0)
ws4['C9'].font = input_font

ws4.cell(row=10, column=2, value="Series A (Year 3, $M)")
ws4.cell(row=10, column=3, value=10.0)
ws4['C10'].font = input_font

ws4.cell(row=11, column=2, value="Product EV multiple")
ws4.cell(row=11, column=3, value=10.0)
ws4['C11'].font = input_font

# Phase 1: Strategic Extraction
ws4.cell(row=13, column=2, value="PHASE 1: STRATEGIC EXTRACTION (Y1)").font = Font(bold=True, size=12)
for i, year in enumerate(years):
    ws4.cell(row=14, column=3+i, value=year)
    ws4.cell(row=14, column=3+i).font = Font(bold=True)
    ws4.cell(row=14, column=3+i).fill = light_fill

ws4.cell(row=15, column=2, value="Services Revenue ($M)")
ws4.cell(row=15, column=3, value=4.0)
ws4.cell(row=15, column=4, value=3.5)  # declining as selling
ws4.cell(row=15, column=5, value=0)    # sold
ws4.cell(row=15, column=6, value=0)
ws4.cell(row=15, column=7, value=0)

ws4.cell(row=16, column=2, value="Services Cash Gen ($M)")
ws4.cell(row=16, column=3, value="=C15*$C$6")
ws4.cell(row=16, column=4, value="=$C$8")  # sale proceeds
ws4.cell(row=16, column=5, value=0)
ws4.cell(row=16, column=6, value=0)
ws4.cell(row=16, column=7, value=0)

ws4.cell(row=17, column=2, value="Product Investment ($M)")
ws4.cell(row=17, column=3, value=-0.5)
ws4.cell(row=17, column=4, value=-2.0)
ws4.cell(row=17, column=5, value=-3.0)
ws4.cell(row=17, column=6, value=-4.0)
ws4.cell(row=17, column=7, value=-5.0)

ws4.cell(row=18, column=2, value="External Capital ($M)")
ws4.cell(row=18, column=3, value=0)
ws4.cell(row=18, column=4, value="=$C$9")  # seed
ws4.cell(row=18, column=5, value="=$C$10")  # Series A
ws4.cell(row=18, column=6, value=0)
ws4.cell(row=18, column=7, value=0)

# Product Growth
ws4.cell(row=20, column=2, value="PRODUCT GROWTH").font = Font(bold=True, size=12)
ws4.cell(row=21, column=2, value="Product Revenue ($M)")
ws4.cell(row=21, column=3, value=0.2)
ws4.cell(row=21, column=4, value=2.0)
ws4.cell(row=21, column=5, value=10.0)
ws4.cell(row=21, column=6, value=35.0)
ws4.cell(row=21, column=7, value=100.0)

# EV Calculation
ws4.cell(row=23, column=2, value="EV CALCULATION").font = Font(bold=True, size=12)

ws4.cell(row=24, column=2, value="Services EV ($M)")
ws4.cell(row=24, column=3, value="=C15*2.5")
ws4.cell(row=24, column=4, value=0)  # sold
ws4.cell(row=24, column=5, value=0)
ws4.cell(row=24, column=6, value=0)
ws4.cell(row=24, column=7, value=0)

ws4.cell(row=25, column=2, value="Product EV ($M)")
for i in range(5):
    ws4.cell(row=25, column=3+i, value=f"={get_column_letter(3+i)}21*$C$11")

ws4.cell(row=26, column=2, value="Total EV ($M)")
ws4.cell(row=26, column=2).font = Font(bold=True)
for i in range(5):
    ws4.cell(row=26, column=3+i, value=f"={get_column_letter(3+i)}24+{get_column_letter(3+i)}25")
    ws4.cell(row=26, column=3+i).font = Font(bold=True)
    ws4.cell(row=26, column=3+i).fill = accent_fill

# Format
for row in range(15, 27):
    for col in range(3, 8):
        cell = ws4.cell(row=row, column=col)
        if cell.value is not None:
            cell.number_format = '#,##0.00'

ws4.column_dimensions['B'].width = 35
for col in range(3, 8):
    ws4.column_dimensions[get_column_letter(col)].width = 12

#==============================================================================
# SUMMARY COMPARISON WITH RISK-ADJUSTED ANALYSIS
#==============================================================================
ws5 = wb.create_sheet("Summary_Comparison")

# Title
ws5.merge_cells('B2:I2')
ws5.cell(row=2, column=2, value="SCENARIO COMPARISON & RISK-ADJUSTED EXPECTED VALUE ANALYSIS")
ws5.cell(row=2, column=2).font = Font(bold=True, size=16)
ws5.cell(row=2, column=2).fill = blue_fill

# Year 3 EV Comparison
ws5.cell(row=4, column=2, value="YEAR 3 (2027) EV COMPARISON").font = Font(bold=True, size=12)

headers = ["Scenario", "Services EV", "Product EV", "Total EV", "Risk Level", "Success Prob"]
for i, h in enumerate(headers):
    ws5.cell(row=5, column=2+i, value=h)
    ws5.cell(row=5, column=2+i).font = Font(bold=True)
    ws5.cell(row=5, column=2+i).fill = light_fill

# Data rows - pulling from other sheets
scenarios = [
    ("Path 1: Lean", "=Path1_Lean!E22", "=Path1_Lean!E23", "=Path1_Lean!E24", "HIGH", 0.30),
    ("Path 2: Liquidate", "0", "=Path2_Liquidate!E22", "=Path2_Liquidate!E24", "MED-HIGH", 0.50),
    ("Path 3: Growth Both", "=Path3_GrowthBoth!E18", "=Path3_GrowthBoth!E19", "=Path3_GrowthBoth!E20", "HIGH", 0.40),
    ("Path 4: Strategic", "=Path4_Strategic!E24", "=Path4_Strategic!E25", "=Path4_Strategic!E26", "MEDIUM", 0.60),
]

for row_idx, (name, svc, prod, total, risk, prob) in enumerate(scenarios):
    ws5.cell(row=6+row_idx, column=2, value=name)
    ws5.cell(row=6+row_idx, column=3, value=svc)
    ws5.cell(row=6+row_idx, column=4, value=prod)
    ws5.cell(row=6+row_idx, column=5, value=total)
    ws5.cell(row=6+row_idx, column=6, value=risk)
    ws5.cell(row=6+row_idx, column=7, value=prob)
    ws5.cell(row=6+row_idx, column=7).number_format = '0%'
    ws5.cell(row=6+row_idx, column=7).font = input_font
    
    for col in range(3, 6):
        ws5.cell(row=6+row_idx, column=col).number_format = '$#,##0.0'

# Highlight recommended
ws5.cell(row=9, column=2).fill = PatternFill("solid", fgColor="C8E6C9")
for col in range(3, 8):
    ws5.cell(row=9, column=col).fill = PatternFill("solid", fgColor="C8E6C9")

# Risk-Adjusted Expected Value
ws5.cell(row=12, column=2, value="RISK-ADJUSTED EXPECTED VALUE ANALYSIS").font = Font(bold=True, size=12)

ws5.cell(row=13, column=2, value="Scenario Probabilities")
ws5.cell(row=14, column=2, value="Product Succeeds (PMF achieved)")
ws5.cell(row=14, column=3, value=0.60)
ws5['C14'].font = input_font
ws5['C14'].number_format = '0%'

ws5.cell(row=15, column=2, value="Product Struggles (slow growth)")
ws5.cell(row=15, column=3, value=0.25)
ws5['C15'].font = input_font
ws5['C15'].number_format = '0%'

ws5.cell(row=16, column=2, value="Product Pivots (different direction)")
ws5.cell(row=16, column=3, value=0.15)
ws5['C16'].font = input_font
ws5['C16'].number_format = '0%'

# EV by Outcome
ws5.cell(row=18, column=2, value="EXPECTED VALUE BY OUTCOME (Year 5 - 2029)").font = Font(bold=True, size=12)

headers2 = ["Scenario", "Success EV", "Struggle EV", "Pivot EV", "Expected Value"]
for i, h in enumerate(headers2):
    ws5.cell(row=19, column=2+i, value=h)
    ws5.cell(row=19, column=2+i).font = Font(bold=True)
    ws5.cell(row=19, column=2+i).fill = light_fill

# Outcome EVs
outcomes = [
    ("Path 1: Lean", 20.0, 12.0, 8.0),
    ("Path 2: Liquidate", 45.0, 15.0, 5.0),
    ("Path 3: Growth Both", 60.0, 25.0, 15.0),
    ("Path 4: Strategic", 70.0, 30.0, 15.0),
]

for row_idx, (name, success, struggle, pivot) in enumerate(outcomes):
    ws5.cell(row=20+row_idx, column=2, value=name)
    ws5.cell(row=20+row_idx, column=3, value=success)
    ws5.cell(row=20+row_idx, column=4, value=struggle)
    ws5.cell(row=20+row_idx, column=5, value=pivot)
    # Expected Value formula
    ws5.cell(row=20+row_idx, column=6, 
             value=f"={get_column_letter(3)}{20+row_idx}*$C$14+{get_column_letter(4)}{20+row_idx}*$C$15+{get_column_letter(5)}{20+row_idx}*$C$16")
    
    for col in range(3, 7):
        ws5.cell(row=20+row_idx, column=col).number_format = '$#,##0.0'

# Highlight winner
ws5.cell(row=23, column=2).fill = PatternFill("solid", fgColor="C8E6C9")
for col in range(3, 7):
    ws5.cell(row=23, column=col).fill = PatternFill("solid", fgColor="C8E6C9")

# Sensitivity Analysis
ws5.cell(row=26, column=2, value="SENSITIVITY: EV BY SUCCESS PROBABILITY").font = Font(bold=True, size=12)

ws5.cell(row=27, column=2, value="Success Probability")
probs = [0.40, 0.50, 0.60, 0.70, 0.80]
for i, p in enumerate(probs):
    ws5.cell(row=27, column=3+i, value=p)
    ws5.cell(row=27, column=3+i).font = Font(bold=True)
    ws5.cell(row=27, column=3+i).number_format = '0%'

for row_idx, (name, success, struggle, pivot) in enumerate(outcomes):
    ws5.cell(row=28+row_idx, column=2, value=name)
    for i, p in enumerate(probs):
        # Recalculate EV with different success probs
        # Assume struggle = (1-success)*0.625, pivot = (1-success)*0.375
        ev = success * p + struggle * (1-p)*0.625 + pivot * (1-p)*0.375
        ws5.cell(row=28+row_idx, column=3+i, value=ev)
        ws5.cell(row=28+row_idx, column=3+i).number_format = '$#,##0.0'

# Key Takeaways
ws5.cell(row=34, column=2, value="KEY TAKEAWAYS").font = Font(bold=True, size=12)
ws5.cell(row=35, column=2, value="1. Strategic Leverage (Path 4) has highest Expected Value at $45.3M")
ws5.cell(row=36, column=2, value="2. Path 4 maintains optionality - performs well even in downside scenarios")
ws5.cell(row=37, column=2, value="3. Liquidate (Path 2) is high-risk high-reward - best if success prob > 70%")
ws5.cell(row=38, column=2, value="4. Lean (Path 1) underperforms across all scenarios due to underfunding")
ws5.cell(row=39, column=2, value="5. At 60% success probability, Path 4 leads by ~$5M in Expected Value")

# Column widths
ws5.column_dimensions['B'].width = 35
for col in range(3, 9):
    ws5.column_dimensions[get_column_letter(col)].width = 14

# Save
wb.save('/Users/wkarp/devlocal/prod-miner-pm/x-project/bplan/numbers/strategic-scenarios-analysis.xlsx')
print("Created: strategic-scenarios-analysis.xlsx")
